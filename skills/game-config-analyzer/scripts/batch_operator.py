#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配表批量操作脚本

对配表进行批量修改
"""

import openpyxl
import json
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Set, Any, Optional
from copy import deepcopy


class BatchOperator:
    """批量操作器"""

    def __init__(self, config_dir: str):
        self.config_dir = Path(config_dir)
        self.backup_dir = self.config_dir / ".backup"
        self.changes_log = []

    def _create_backup(self):
        """创建备份"""
        if not self.backup_dir.exists():
            self.backup_dir.mkdir(parents=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / timestamp

        if backup_path.exists():
            return backup_path

        shutil.copytree(self.config_dir, backup_path)
        print(f"已创建备份: {backup_path}")

        return backup_path

    def add_field(self, tables: List[str], field_def: Dict) -> Dict[str, Any]:
        """批量添加字段"""
        print(f"为 {len(tables)} 个表添加字段: {field_def['name']}")

        # 创建备份
        backup_path = self._create_backup()

        results = {
            "operation": "add_field",
            "field": field_def,
            "success": [],
            "failed": []
        }

        for table_name in tables:
            file_path = self._find_table_file(table_name)
            if not file_path:
                results["failed"].append({"table": table_name, "reason": "文件未找到"})
                continue

            try:
                wb = openpyxl.load_workbook(file_path)

                for sheet in wb.worksheets:
                    if self._should_process_sheet(sheet, table_name):
                        # 添加新列
                        new_col = sheet.max_column + 1

                        # 填充表头
                        sheet.cell(1, new_col, field_def.get("chs_name", field_def["name"]))
                        sheet.cell(2, new_col, field_def.get("type", "string"))
                        sheet.cell(3, new_col, field_def["name"])
                        sheet.cell(4, new_col, field_def.get("export", "server"))

                        # 填充默认值
                        if "default" in field_def:
                            for row in range(5, sheet.max_row + 1):
                                sheet.cell(row, new_col, field_def["default"])

                wb.save(file_path)
                results["success"].append(table_name)
                print(f"  ✓ {table_name}")

                # 记录变更
                self.changes_log.append({
                    "table": table_name,
                    "operation": "add_field",
                    "field": field_def["name"],
                    "value": field_def
                })

            except Exception as e:
                results["failed"].append({"table": table_name, "reason": str(e)})
                print(f"  ✗ {table_name}: {e}")

        return results

    def remove_field(self, tables: List[str], field_name: str) -> Dict[str, Any]:
        """批量删除字段"""
        print(f"从 {len(tables)} 个表删除字段: {field_name}")

        backup_path = self._create_backup()

        results = {
            "operation": "remove_field",
            "field": field_name,
            "success": [],
            "failed": []
        }

        for table_name in tables:
            file_path = self._find_table_file(table_name)
            if not file_path:
                results["failed"].append({"table": table_name, "reason": "文件未找到"})
                continue

            try:
                wb = openpyxl.load_workbook(file_path)

                for sheet in wb.worksheets:
                    if self._should_process_sheet(sheet, table_name):
                        # 找到字段所在的列
                        col_to_delete = None
                        for col in range(1, sheet.max_column + 1):
                            if sheet.cell(3, col).value == field_name:
                                col_to_delete = col
                                break

                        if col_to_delete:
                            sheet.delete_cols(col_to_delete)

                wb.save(file_path)
                results["success"].append(table_name)
                print(f"  ✓ {table_name}")

                self.changes_log.append({
                    "table": table_name,
                    "operation": "remove_field",
                    "field": field_name
                })

            except Exception as e:
                results["failed"].append({"table": table_name, "reason": str(e)})
                print(f"  ✗ {table_name}: {e}")

        return results

    def update_values(self, table: str, condition: Dict, update: Dict) -> Dict[str, Any]:
        """按条件修改值"""
        print(f"更新 {table} 表的值")

        backup_path = self._create_backup()

        results = {
            "operation": "update_values",
            "table": table,
            "condition": condition,
            "update": update,
            "updated": 0,
            "failed": 0
        }

        file_path = self._find_table_file(table)
        if not file_path:
            results["failed"] = 1
            return results

        try:
            wb = openpyxl.load_workbook(file_path)

            for sheet in wb.worksheets:
                if self._should_process_sheet(sheet, table):
                    # 获取表头
                    headers = {}
                    for col in range(1, sheet.max_column + 1):
                        field_name = sheet.cell(3, col).value
                        if field_name:
                            headers[str(field_name)] = col

                    # 遍历数据行
                    for row_idx in range(5, sheet.max_row + 1):
                        # 检查条件
                        match = True
                        for field, expected_value in condition.items():
                            if field in headers:
                                actual_value = sheet.cell(row_idx, headers[field]).value
                                if actual_value != expected_value:
                                    match = False
                                    break

                        # 执行更新
                        if match:
                            for field, new_value in update.items():
                                if field in headers:
                                    old_value = sheet.cell(row_idx, headers[field]).value
                                    sheet.cell(row_idx, headers[field]).value = new_value
                                    results["updated"] += 1

                                    self.changes_log.append({
                                        "table": table,
                                        "operation": "update_value",
                                        "row": row_idx,
                                        "field": field,
                                        "old_value": old_value,
                                        "new_value": new_value
                                    })

            wb.save(file_path)
            print(f"  更新了 {results['updated']} 个单元格")

        except Exception as e:
            results["failed"] = 1
            print(f"  ✗ 错误: {e}")

        return results

    def replace_values(self, table: str, field: str, old_value: Any, new_value: Any) -> Dict[str, Any]:
        """批量替换值"""
        print(f"在 {table}.{field} 中替换 '{old_value}' → '{new_value}'")

        backup_path = self._create_backup()

        results = {
            "operation": "replace_values",
            "table": table,
            "field": field,
            "old_value": old_value,
            "new_value": new_value,
            "replaced": 0
        }

        file_path = self._find_table_file(table)
        if not file_path:
            return results

        try:
            wb = openpyxl.load_workbook(file_path)

            for sheet in wb.worksheets:
                if self._should_process_sheet(sheet, table):
                    # 找到字段列
                    field_col = None
                    for col in range(1, sheet.max_column + 1):
                        if sheet.cell(3, col).value == field:
                            field_col = col
                            break

                    if field_col:
                        # 替换值
                        for row_idx in range(5, sheet.max_row + 1):
                            cell_value = sheet.cell(row_idx, field_col).value
                            if cell_value == old_value:
                                sheet.cell(row_idx, field_col).value = new_value
                                results["replaced"] += 1

                                self.changes_log.append({
                                    "table": table,
                                    "operation": "replace_value",
                                    "row": row_idx,
                                    "field": field,
                                    "old_value": old_value,
                                    "new_value": new_value
                                })

            wb.save(file_path)
            print(f"  替换了 {results['replaced']} 处")

        except Exception as e:
            print(f"  ✗ 错误: {e}")

        return results

    def preview_operation(self, operation: Dict) -> Dict[str, Any]:
        """预览操作结果（不实际修改）"""
        print("预览操作（不实际修改）")

        operation_type = operation.get("type")

        if operation_type == "add_field":
            tables = operation.get("tables", [])
            field_def = operation.get("field_def", {})
            return {
                "operation": "preview_add_field",
                "tables": tables,
                "field": field_def,
                "affected_tables": len(tables),
                "note": "将在每个表的末尾添加新列"
            }

        elif operation_type == "update_values":
            table = operation.get("table")
            condition = operation.get("condition", {})
            update = operation.get("update", {})

            # 查找匹配的行
            file_path = self._find_table_file(table)
            if not file_path:
                return {"error": "表未找到"}

            matches = self._find_matching_rows(file_path, table, condition)
            return {
                "operation": "preview_update_values",
                "table": table,
                "condition": condition,
                "update": update,
                "matching_rows": len(matches),
                "changes": len(matches) * len(update)
            }

        return {"error": "未知操作类型"}

    def _find_table_file(self, table_name: str) -> Optional[Path]:
        """查找表文件"""
        # 尝试直接匹配
        direct_file = self.config_dir / f"{table_name}.xlsx"
        if direct_file.exists():
            return direct_file

        # 递归搜索
        for file_path in self.config_dir.rglob("*.xlsx"):
            if file_path.stem == table_name or file_path.name == table_name:
                return file_path

        return None

    def _should_process_sheet(self, sheet, table_name: str) -> bool:
        """判断是否应该处理该 sheet"""
        # 简化处理：如果表名和 sheet 名匹配就处理
        return sheet.title.lower() == table_name.lower() or table_name.lower() in str(sheet.title).lower()

    def _find_matching_rows(self, file_path: Path, table_name: str, condition: Dict) -> List[Dict]:
        """查找匹配条件的行"""
        matches = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet in wb.worksheets:
                if self._should_process_sheet(sheet, table_name):
                    # 获取表头
                    headers = {}
                    for col in range(1, sheet.max_column + 1):
                        field_name = sheet.cell(3, col).value
                        if field_name:
                            headers[str(field_name)] = col

                    # 检查每行
                    for row_idx in range(5, sheet.max_row + 1):
                        match = True
                        row_data = {}

                        for field, expected_value in condition.items():
                            if field in headers:
                                actual_value = sheet.cell(row_idx, headers[field]).value
                                if actual_value != expected_value:
                                    match = False
                                    break
                                row_data[field] = actual_value

                        if match:
                            matches.append({"row": row_idx, "data": row_data})

        except Exception as e:
            print(f"查找匹配行时出错: {e}")

        return matches

    def rollback(self, backup_path: str = None):
        """回滚操作"""
        if backup_path is None:
            # 使用最新的备份
            if not self.backup_dir.exists():
                print("没有找到备份目录")
                return

            backups = sorted(self.backup_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True)
            if not backups:
                print("没有找到备份")
                return

            backup_path = backups[0]

        print(f"从备份恢复: {backup_path}")

        # 删除当前内容
        for item in self.config_dir.iterdir():
            if item.name != ".backup":
                if item.is_dir():
                    shutil.rmtree(item)
                else:
                    item.unlink()

        # 恢复备份
        shutil.copytree(backup_path, self.config_dir, dirs_exist_ok=True)
        print("恢复完成")

    def save_changes_log(self, log_file: str = "batch_changes.json"):
        """保存变更日志"""
        log_data = {
            "timestamp": datetime.now().isoformat(),
            "changes": self.changes_log
        }

        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(log_data, f, ensure_ascii=False, indent=2)

        print(f"变更日志已保存到: {log_file}")

    def generate_report(self, results: Dict) -> str:
        """生成操作报告"""
        lines = [
            "# 批量操作报告",
            f"",
            f"**操作**: {results.get('operation', 'Unknown')}",
            f"**时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f""
        ]

        operation = results.get("operation")

        if operation == "add_field":
            field = results["field"]
            lines.extend([
                f"## 添加字段",
                f"",
                f"- 字段名: {field.get('name')}",
                f"- 类型: {field.get('type')}",
                f"- 中文名: {field.get('chs_name', '-')}",
                f"- 默认值: {field.get('default', '-')}",
                f""
            ])

            if results["success"]:
                lines.extend([
                    f"### 成功 ({len(results['success'])} 个表)",
                    f"",
                    *[f"- ✓ {table}" for table in results["success"]],
                    f""
                ])

            if results["failed"]:
                lines.extend([
                    f"### 失败 ({len(results['failed'])} 个表)",
                    f"",
                ])
                for fail in results["failed"]:
                    lines.append(f"- ✗ {fail['table']}: {fail['reason']}")
                lines.append("")

        elif operation == "update_values":
            lines.extend([
                f"## 更新值",
                f"",
                f"- 表: {results['table']}",
                f"- 条件: {results['condition']}",
                f"- 更新: {results['update']}",
                f"- 更新单元格数: {results['updated']}",
                f""
            ])

        return "\n".join(lines)


def main():
    """主函数"""
    import sys

    if len(sys.argv) < 2:
        print("用法:")
        print("  python batch_operator.py <配表目录> <操作> [参数...]")
        print("")
        print("操作:")
        print("  add <表列表> <字段定义.json>  - 添加字段")
        print("  remove <表列表> <字段名>      - 删除字段")
        print("  update <表名> <条件.json> <更新.json> - 更新值")
        print("  preview <操作.json>           - 预览操作")
        print("")
        print("示例:")
        print("  python batch_operator.py ./configs add Hero,Card field_def.json")
        print("  python batch_operator.py ./configs remove Hero TempField")
        sys.exit(1)

    config_dir = sys.argv[1]
    operation = sys.argv[2] if len(sys.argv) > 2 else ""

    operator = BatchOperator(config_dir)

    if operation == "add":
        tables = sys.argv[3].split(",") if len(sys.argv) > 3 else []
        field_def_file = sys.argv[4] if len(sys.argv) > 4 else "field_def.json"

        with open(field_def_file, 'r', encoding='utf-8') as f:
            field_def = json.load(f)

        results = operator.add_field(tables, field_def)

    elif operation == "remove":
        tables = sys.argv[3].split(",") if len(sys.argv) > 3 else []
        field_name = sys.argv[4] if len(sys.argv) > 4 else ""

        results = operator.remove_field(tables, field_name)

    elif operation == "update":
        table = sys.argv[3] if len(sys.argv) > 3 else ""
        condition_file = sys.argv[4] if len(sys.argv) > 4 else "condition.json"
        update_file = sys.argv[5] if len(sys.argv) > 5 else "update.json"

        with open(condition_file, 'r', encoding='utf-8') as f:
            condition = json.load(f)
        with open(update_file, 'r', encoding='utf-8') as f:
            update = json.load(f)

        results = operator.update_values(table, condition, update)

    elif operation == "preview":
        op_file = sys.argv[3] if len(sys.argv) > 3 else "operation.json"

        with open(op_file, 'r', encoding='utf-8') as f:
            operation_def = json.load(f)

        results = operator.preview_operation(operation_def)

    else:
        print(f"未知操作: {operation}")
        sys.exit(1)

    # 保存结果
    output_file = "batch_operation_result.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"结果已保存到: {output_file}")

    # 生成报告
    report = operator.generate_report(results)
    report_file = output_file.replace('.json', '_report.md')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"报告已保存到: {report_file}")

    # 保存变更日志
    operator.save_changes_log()


if __name__ == "__main__":
    main()
