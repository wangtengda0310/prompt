#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配表差异分析脚本

对比两个版本的配表，识别变更
"""

import openpyxl
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any
from collections import defaultdict


class DiffAnalyzer:
    """差异分析器"""

    def __init__(self, dir_a: str, dir_b: str):
        self.dir_a = Path(dir_a)
        self.dir_b = Path(dir_b)

    def compare(self) -> Dict[str, Any]:
        """对比两个版本"""
        # 分析两个版本
        analysis_a = self._analyze_directory(self.dir_a, "version_a")
        analysis_b = self._analyze_directory(self.dir_b, "version_b")

        # 执行对比
        result = {
            "timestamp": datetime.now().isoformat(),
            "version_a": str(self.dir_a),
            "version_b": str(self.dir_b),
            "summary": self._generate_summary(analysis_a, analysis_b),
            "tables": self._compare_tables(analysis_a, analysis_b),
            "structures": self._compare_structures(analysis_a, analysis_b),
            "data": self._compare_data(analysis_a, analysis_b),
            "constraints": self._compare_constraints(analysis_a, analysis_b)
        }

        return result

    def _analyze_directory(self, directory: Path, version: str) -> Dict:
        """分析目录中的配表"""
        tables = {}
        for file_path in directory.rglob("*.xlsx"):
            if "~$" in file_path.name:
                continue

            try:
                file_info = self._analyze_file(file_path)
                if file_info:
                    tables[file_info["filename"]] = file_info
            except Exception as e:
                print(f"警告: 无法读取 {file_path}: {e}")

        return {"version": version, "tables": tables}

    def _analyze_file(self, file_path: Path) -> Dict:
        """分析单个文件"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = []

        for sheet in wb.worksheets:
            sheet_info = self._analyze_sheet(sheet)
            if sheet_info:
                sheets.append(sheet_info)

        return {
            "filename": file_path.name,
            "path": str(file_path),
            "sheets": sheets
        }

    def _analyze_sheet(self, sheet) -> Dict:
        """分析 Sheet"""
        # 假设标准 4 行表头
        header_row = 3
        data_start_row = 4

        headers = []
        for col in range(1, sheet.max_column + 1):
            field_name = sheet.cell(header_row, col).value
            if field_name:
                headers.append({
                    "column": col,
                    "name": str(field_name),
                    "type": str(sheet.cell(2, col).value) if sheet.cell(2, col).value else ""
                })

        # 收集数据（用于数据对比）
        data = []
        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)

        return {
            "name": sheet.title,
            "headers": headers,
            "row_count": len(data),
            "data_sample": data[:10]  # 只保存样本数据
        }

    def _generate_summary(self, analysis_a: Dict, analysis_b: Dict) -> Dict:
        """生成变更摘要"""
        tables_a = set(analysis_a["tables"].keys())
        tables_b = set(analysis_b["tables"].keys())

        return {
            "tables_added": list(tables_b - tables_a),
            "tables_removed": list(tables_a - tables_b),
            "tables_modified": len(tables_a & tables_b),
            "total_changes": 0  # 将在后续计算
        }

    def _compare_tables(self, analysis_a: Dict, analysis_b: Dict) -> List[Dict]:
        """对比表列表"""
        changes = []
        summary = self._generate_summary(analysis_a, analysis_b)

        # 新增表
        for table in summary["tables_added"]:
            changes.append({
                "type": "table_added",
                "table": table,
                "details": analysis_b["tables"][table]
            })

        # 删除表
        for table in summary["tables_removed"]:
            changes.append({
                "type": "table_removed",
                "table": table,
                "details": analysis_a["tables"][table]
            })

        return changes

    def _compare_structures(self, analysis_a: Dict, analysis_b: Dict) -> List[Dict]:
        """对比表结构"""
        changes = []
        common_tables = set(analysis_a["tables"].keys()) & set(analysis_b["tables"].keys())

        for table_name in common_tables:
            table_a = analysis_a["tables"][table_name]
            table_b = analysis_b["tables"][table_name]

            # 对比每个 sheet
            for sheet_a in table_a["sheets"]:
                sheet_b = self._find_sheet(table_b["sheets"], sheet_a["name"])
                if not sheet_b:
                    changes.append({
                        "type": "sheet_removed",
                        "table": table_name,
                        "sheet": sheet_a["name"]
                    })
                    continue

                # 对比字段
                headers_a = {h["name"]: h for h in sheet_a["headers"]}
                headers_b = {h["name"]: h for h in sheet_b["headers"]}

                # 新增字段
                for field_name in set(headers_b.keys()) - set(headers_a.keys()):
                    changes.append({
                        "type": "field_added",
                        "table": table_name,
                        "sheet": sheet_a["name"],
                        "field": field_name,
                        "field_type": headers_b[field_name]["type"]
                    })

                # 删除字段
                for field_name in set(headers_a.keys()) - set(headers_b.keys()):
                    changes.append({
                        "type": "field_removed",
                        "table": table_name,
                        "sheet": sheet_a["name"],
                        "field": field_name
                    })

        return changes

    def _compare_data(self, analysis_a: Dict, analysis_b: Dict) -> List[Dict]:
        """对比数据变更"""
        changes = []
        common_tables = set(analysis_a["tables"].keys()) & set(analysis_b["tables"].keys())

        for table_name in common_tables:
            table_a = analysis_a["tables"][table_name]
            table_b = analysis_b["tables"][table_name]

            for sheet_a in table_a["sheets"]:
                sheet_b = self._find_sheet(table_b["sheets"], sheet_a["name"])
                if not sheet_b:
                    continue

                # 检查行数变化
                if sheet_a["row_count"] != sheet_b["row_count"]:
                    changes.append({
                        "type": "row_count_changed",
                        "table": table_name,
                        "sheet": sheet_a["name"],
                        "old_count": sheet_a["row_count"],
                        "new_count": sheet_b["row_count"]
                    })

        return changes

    def _compare_constraints(self, analysis_a: Dict, analysis_b: Dict) -> List[Dict]:
        """对比约束规则"""
        # 这里简化处理，实际需要更复杂的约束提取和对比
        return []

    def _find_sheet(self, sheets: List[Dict], name: str) -> Dict:
        """查找指定名称的 sheet"""
        for sheet in sheets:
            if sheet["name"] == name:
                return sheet
        return None

    def generate_report(self, diff_result: Dict) -> str:
        """生成差异报告"""
        lines = [
            "# 配表差异分析报告",
            f"",
            f"**生成时间**: {diff_result['timestamp']}",
            f"",
            f"## 变更摘要",
            f"",
            f"- 新增表: {len(diff_result['summary']['tables_added'])} 个",
            f"- 删除表: {len(diff_result['summary']['tables_removed'])} 个",
            f"- 修改表: {diff_result['summary']['tables_modified']} 个",
            f""
        ]

        # 表级变更
        if diff_result["tables"]:
            lines.extend([
                f"## 表级变更",
                f""
            ])
            for change in diff_result["tables"]:
                if change["type"] == "table_added":
                    lines.append(f"- ✅ 新增表: {change['table']}")
                elif change["type"] == "table_removed":
                    lines.append(f"- ❌ 删除表: {change['table']}")
            lines.append("")

        # 结构变更
        if diff_result["structures"]:
            lines.extend([
                f"## 结构变更",
                f""
            ])
            for change in diff_result["structures"]:
                if change["type"] == "field_added":
                    lines.append(f"- ➕ **{change['table']}.{change['sheet']}**: 新增字段 `{change['field']}` ({change['field_type']})")
                elif change["type"] == "field_removed":
                    lines.append(f"- ➖ **{change['table']}.{change['sheet']}**: 删除字段 `{change['field']}`")
            lines.append("")

        # 数据变更
        if diff_result["data"]:
            lines.extend([
                f"## 数据变更",
                f""
            ])
            for change in diff_result["data"]:
                if change["type"] == "row_count_changed":
                    lines.append(f"- 📊 **{change['table']}.{change['sheet']}**: 行数 {change['old_count']} → {change['new_count']}")
            lines.append("")

        return "\n".join(lines)


def main():
    """主函数"""
    import sys

    if len(sys.argv) < 3:
        print("用法: python diff_analyzer.py <目录A> <目录B> [输出文件]")
        sys.exit(1)

    dir_a = sys.argv[1]
    dir_b = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else "diff_result.json"

    # 执行对比
    print(f"对比目录:")
    print(f"  A: {dir_a}")
    print(f"  B: {dir_b}")
    print()

    analyzer = DiffAnalyzer(dir_a, dir_b)
    result = analyzer.compare()

    # 保存 JSON 结果
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"JSON 结果已保存到: {output_file}")

    # 生成 Markdown 报告
    report = analyzer.generate_report(result)
    report_file = output_file.replace('.json', '_report.md')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"Markdown 报告已保存到: {report_file}")

    # 打印摘要
    print()
    print("变更摘要:")
    print(f"  新增表: {len(result['summary']['tables_added'])} 个")
    print(f"  删除表: {len(result['summary']['tables_removed'])} 个")
    print(f"  修改表: {result['summary']['tables_modified']} 个")
    print(f"  结构变更: {len(result['structures'])} 项")
    print(f"  数据变更: {len(result['data'])} 项")


if __name__ == "__main__":
    main()
