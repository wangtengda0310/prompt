#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配表分析 Subagent 调度器

用于将大型分析任务拆分为多个并行子任务，提高分析效率。
"""

import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field
from enum import Enum
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time


class SubagentTaskType(Enum):
    """子任务类型"""
    SCAN = "scan"                     # 配表扫描
    STRUCTURE_ANALYSIS = "structure"  # 结构分析
    RELATION_ANALYSIS = "relation"    # 关系分析
    CONSTRAINT_EXTRACTION = "constraint"  # 约束提取
    ANTI_PATTERN = "antipattern"      # 反模式检测
    TIME_CONSTRAINT = "time"          # 时间约束分析
    DATA_QUALITY = "quality"          # 数据质量检查
    ENUM_ANALYSIS = "enum"            # 枚举分析
    DOC_GENERATION = "doc"            # 文档生成


@dataclass
class SubagentTask:
    """子任务定义"""
    task_id: str
    task_type: SubagentTaskType
    description: str
    params: Dict[str, Any] = field(default_factory=dict)
    dependencies: List[str] = field(default_factory=list)
    priority: int = 0  # 越小优先级越高
    timeout: int = 120  # 超时时间（秒）
    status: str = "pending"  # pending, running, completed, failed
    result: Optional[Dict] = None
    error: Optional[str] = None


@dataclass
class AnalysisPlan:
    """分析计划"""
    plan_id: str
    config_dir: str
    output_dir: str
    tasks: List[SubagentTask] = field(default_factory=list)
    parallel_groups: List[List[str]] = field(default_factory=list)

    def get_task_by_id(self, task_id: str) -> Optional[SubagentTask]:
        for task in self.tasks:
            if task.task_id == task_id:
                return task
        return None


class SubagentScheduler:
    """Subagent 调度器"""

    def __init__(self, config_dir: str, output_dir: str = None, max_workers: int = 4):
        self.config_dir = Path(config_dir)
        self.output_dir = Path(output_dir) if output_dir else self.config_dir
        self.max_workers = max_workers
        self.results: Dict[str, Any] = {}
        self._lock = threading.Lock()

    def create_full_analysis_plan(self) -> AnalysisPlan:
        """创建完整分析计划"""
        plan = AnalysisPlan(
            plan_id=f"analysis_{int(time.time())}",
            config_dir=str(self.config_dir),
            output_dir=str(self.output_dir)
        )

        # 第1阶段：独立扫描任务（可并行）
        plan.tasks.append(SubagentTask(
            task_id="scan_core",
            task_type=SubagentTaskType.SCAN,
            description="扫描核心配置表（Hero/Card/Skill等）",
            params={"categories": ["hero", "card", "skill", "item", "buff"]},
            priority=0
        ))

        plan.tasks.append(SubagentTask(
            task_id="scan_enum",
            task_type=SubagentTaskType.SCAN,
            description="扫描枚举表",
            params={"categories": ["enum"], "subdir": "enum"},
            priority=0
        ))

        plan.tasks.append(SubagentTask(
            task_id="scan_activity",
            task_type=SubagentTaskType.SCAN,
            description="扫描活动相关表",
            params={"categories": ["activity", "task", "achieve"]},
            priority=0
        ))

        plan.tasks.append(SubagentTask(
            task_id="scan_pve",
            task_type=SubagentTaskType.SCAN,
            description="扫描PVE相关表",
            params={"categories": ["pve", "guild", "arena"]},
            priority=0
        ))

        # 第2阶段：依赖扫描结果的分析任务（可并行）
        plan.tasks.append(SubagentTask(
            task_id="relation_analysis",
            task_type=SubagentTaskType.RELATION_ANALYSIS,
            description="分析表间引用关系",
            dependencies=["scan_core", "scan_enum"],
            priority=1
        ))

        plan.tasks.append(SubagentTask(
            task_id="constraint_extraction",
            task_type=SubagentTaskType.CONSTRAINT_EXTRACTION,
            description="提取字段约束规则",
            dependencies=["scan_core"],
            priority=1
        ))

        plan.tasks.append(SubagentTask(
            task_id="enum_analysis",
            task_type=SubagentTaskType.ENUM_ANALYSIS,
            description="分析枚举类型定义",
            dependencies=["scan_enum"],
            priority=1
        ))

        # 第3阶段：深度分析（可并行）
        plan.tasks.append(SubagentTask(
            task_id="anti_pattern",
            task_type=SubagentTaskType.ANTI_PATTERN,
            description="检测设计反模式",
            dependencies=["relation_analysis"],
            priority=2
        ))

        plan.tasks.append(SubagentTask(
            task_id="time_constraint",
            task_type=SubagentTaskType.TIME_CONSTRAINT,
            description="提取时间约束规则",
            dependencies=["constraint_extraction"],
            priority=2
        ))

        plan.tasks.append(SubagentTask(
            task_id="data_quality",
            task_type=SubagentTaskType.DATA_QUALITY,
            description="数据质量检查",
            dependencies=["scan_core", "scan_enum"],
            priority=2
        ))

        # 第4阶段：文档生成（依赖所有分析任务）
        plan.tasks.append(SubagentTask(
            task_id="doc_generation",
            task_type=SubagentTaskType.DOC_GENERATION,
            description="生成分析报告",
            dependencies=["relation_analysis", "constraint_extraction",
                         "anti_pattern", "time_constraint", "data_quality"],
            priority=3
        ))

        # 计算并行组
        plan.parallel_groups = self._compute_parallel_groups(plan.tasks)

        return plan

    def _compute_parallel_groups(self, tasks: List[SubagentTask]) -> List[List[str]]:
        """计算可并行执行的任务组"""
        groups = []
        remaining = {t.task_id: t for t in tasks}
        completed = set()

        while remaining:
            # 找出所有依赖已满足的任务
            ready = []
            for task_id, task in remaining.items():
                if all(dep in completed for dep in task.dependencies):
                    ready.append((task.priority, task_id))

            if not ready:
                # 存在循环依赖
                break

            # 按优先级排序
            ready.sort()
            group = [task_id for _, task_id in ready]
            groups.append(group)

            # 标记为完成
            for task_id in group:
                completed.add(task_id)
                del remaining[task_id]

        return groups

    def execute_task(self, task: SubagentTask) -> Dict[str, Any]:
        """执行单个任务"""
        task.status = "running"

        try:
            # 根据任务类型执行不同的分析
            if task.task_type == SubagentTaskType.SCAN:
                result = self._execute_scan(task)
            elif task.task_type == SubagentTaskType.RELATION_ANALYSIS:
                result = self._execute_relation_analysis(task)
            elif task.task_type == SubagentTaskType.CONSTRAINT_EXTRACTION:
                result = self._execute_constraint_extraction(task)
            elif task.task_type == SubagentTaskType.ENUM_ANALYSIS:
                result = self._execute_enum_analysis(task)
            elif task.task_type == SubagentTaskType.ANTI_PATTERN:
                result = self._execute_anti_pattern(task)
            elif task.task_type == SubagentTaskType.TIME_CONSTRAINT:
                result = self._execute_time_constraint(task)
            elif task.task_type == SubagentTaskType.DATA_QUALITY:
                result = self._execute_data_quality(task)
            elif task.task_type == SubagentTaskType.DOC_GENERATION:
                result = self._execute_doc_generation(task)
            else:
                result = {"error": f"未知任务类型: {task.task_type}"}

            task.result = result
            task.status = "completed"

        except Exception as e:
            task.error = str(e)
            task.status = "failed"
            result = {"error": str(e)}

        with self._lock:
            self.results[task.task_id] = result

        return result

    def _execute_scan(self, task: SubagentTask) -> Dict[str, Any]:
        """执行扫描任务"""
        import openpyxl

        categories = task.params.get("categories", [])
        subdir = task.params.get("subdir", "")

        scan_dir = self.config_dir / subdir if subdir else self.config_dir

        result = {
            "task_id": task.task_id,
            "files": [],
            "stats": {"total": 0, "sheets": 0, "rows": 0}
        }

        for xlsx_file in scan_dir.rglob("*.xlsx"):
            if "~$" in xlsx_file.name:
                continue

            # 检查是否符合分类
            file_stem = xlsx_file.stem.lower()
            if categories:
                matched = any(cat.lower() in file_stem for cat in categories)
                if not matched:
                    continue

            try:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True, read_only=True)

                file_info = {
                    "name": xlsx_file.name,
                    "path": str(xlsx_file.relative_to(self.config_dir)),
                    "sheets": []
                }

                for sheet in wb.worksheets:
                    rows = sum(1 for row in sheet.iter_rows(min_row=5, values_only=True)
                              if any(cell is not None for cell in row))

                    file_info["sheets"].append({
                        "name": sheet.title,
                        "rows": rows
                    })
                    result["stats"]["rows"] += rows
                    result["stats"]["sheets"] += 1

                result["files"].append(file_info)
                result["stats"]["total"] += 1

                wb.close()

            except Exception as e:
                pass

        return result

    def _execute_relation_analysis(self, task: SubagentTask) -> Dict[str, Any]:
        """执行关系分析"""
        import openpyxl
        import re

        scan_result = self.results.get("scan_core", {})
        enum_result = self.results.get("scan_enum", {})

        result = {
            "task_id": task.task_id,
            "relations": [],
            "center_tables": [],
            "orphan_tables": []
        }

        # 构建表名映射
        table_map = {}
        for file_info in scan_result.get("files", []):
            for sheet in file_info.get("sheets", []):
                table_map[sheet["name"]] = file_info["name"]

        # 分析引用关系
        ref_counts = {}
        for file_info in scan_result.get("files", []):
            file_path = self.config_dir / file_info["path"]
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
                for sheet in wb.worksheets:
                    # 读取第3行（字段名）
                    field_names = []
                    for cell in sheet[3]:
                        if cell.value:
                            field_names.append(str(cell.value))

                    # 分析 ID 引用
                    for field in field_names:
                        id_match = re.match(r'(.+)Id$', field)
                        if id_match:
                            ref_table = id_match.group(1)
                            if ref_table in table_map:
                                result["relations"].append({
                                    "source": sheet.title,
                                    "source_field": field,
                                    "target": ref_table,
                                    "type": "id_reference"
                                })
                                ref_counts[ref_table] = ref_counts.get(ref_table, 0) + 1
                wb.close()
            except:
                pass

        # 识别中心表
        sorted_refs = sorted(ref_counts.items(), key=lambda x: x[1], reverse=True)
        result["center_tables"] = [{"table": t, "refs": c} for t, c in sorted_refs[:10]]

        # 识别孤立表
        referenced = set(r["target"] for r in result["relations"])
        referencing = set(r["source"] for r in result["relations"])
        all_tables = set(table_map.keys())
        result["orphan_tables"] = list(all_tables - referenced - referencing)

        return result

    def _execute_constraint_extraction(self, task: SubagentTask) -> Dict[str, Any]:
        """执行约束提取"""
        import openpyxl

        scan_result = self.results.get("scan_core", {})

        result = {
            "task_id": task.task_id,
            "constraints": {
                "time_fields": [],
                "id_references": [],
                "enum_fields": [],
                "required_fields": []
            }
        }

        for file_info in scan_result.get("files", []):
            file_path = self.config_dir / file_info["path"]
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
                for sheet in wb.worksheets:
                    # 读取表头
                    row1 = [cell.value for cell in sheet[1]]  # 中文名
                    row2 = [cell.value for cell in sheet[2]]  # 类型
                    row3 = [cell.value for cell in sheet[3]]  # 字段名
                    row4 = [cell.value for cell in sheet[4]]  # 导出标识

                    for i, (chs, ftype, fname, export) in enumerate(zip(row1, row2, row3, row4)):
                        if not fname:
                            continue

                        fname_str = str(fname)
                        ftype_str = str(ftype) if ftype else ""

                        # 时间字段
                        time_keywords = ["time", "date", "start", "end", "expire", "valid"]
                        if any(kw in fname_str.lower() for kw in time_keywords):
                            result["constraints"]["time_fields"].append({
                                "table": sheet.title,
                                "field": fname_str,
                                "chs_name": chs,
                                "type": ftype_str
                            })

                        # ID 引用
                        if fname_str.endswith("Id") and fname_str != "Id":
                            result["constraints"]["id_references"].append({
                                "table": sheet.title,
                                "field": fname_str,
                                "target": fname_str[:-2]
                            })

                        # 枚举字段
                        if ftype_str.startswith("E") and not ftype_str.startswith("int"):
                            result["constraints"]["enum_fields"].append({
                                "table": sheet.title,
                                "field": fname_str,
                                "enum_type": ftype_str
                            })

                        # 必填字段
                        export_str = str(export) if export else ""
                        if "server" in export_str or "client" in export_str:
                            result["constraints"]["required_fields"].append({
                                "table": sheet.title,
                                "field": fname_str,
                                "export": export_str
                            })

                wb.close()
            except:
                pass

        return result

    def _execute_enum_analysis(self, task: SubagentTask) -> Dict[str, Any]:
        """执行枚举分析"""
        import openpyxl

        scan_result = self.results.get("scan_enum", {})

        result = {
            "task_id": task.task_id,
            "enums": []
        }

        for file_info in scan_result.get("files", []):
            file_path = self.config_dir / file_info["path"]
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                for sheet in wb.worksheets:
                    enum_info = {
                        "name": sheet.title,
                        "file": file_info["name"],
                        "values": []
                    }

                    for row in sheet.iter_rows(min_row=5, values_only=True):
                        if row[0] and row[1]:
                            enum_info["values"].append({
                                "name": str(row[0]),
                                "value": row[1],
                                "desc": str(row[2]) if len(row) > 2 and row[2] else ""
                            })

                    if enum_info["values"]:
                        result["enums"].append(enum_info)

                wb.close()
            except:
                pass

        return result

    def _execute_anti_pattern(self, task: SubagentTask) -> Dict[str, Any]:
        """执行反模式检测"""
        relation_result = self.results.get("relation_analysis", {})

        result = {
            "task_id": task.task_id,
            "circular_refs": [],
            "orphan_tables": [],
            "over_depended": []
        }

        # 检测循环引用
        relations = relation_result.get("relations", [])
        graph = {}
        for rel in relations:
            source = rel["source"]
            target = rel["target"]
            if source not in graph:
                graph[source] = []
            graph[source].append(target)

        # 简单的循环检测
        for node in graph:
            visited = set()
            path = []
            if self._has_cycle(graph, node, visited, path):
                result["circular_refs"].append({
                    "start": node,
                    "path": path
                })

        # 孤立表
        result["orphan_tables"] = relation_result.get("orphan_tables", [])

        # 过度依赖（引用超过5个表）
        ref_counts = {}
        for rel in relations:
            source = rel["source"]
            ref_counts[source] = ref_counts.get(source, 0) + 1

        for table, count in ref_counts.items():
            if count > 5:
                result["over_depended"].append({
                    "table": table,
                    "ref_count": count
                })

        return result

    def _has_cycle(self, graph: Dict, node: str, visited: set, path: List) -> bool:
        """检测图中是否有循环"""
        if node in visited:
            return True
        visited.add(node)
        path.append(node)

        for neighbor in graph.get(node, []):
            if self._has_cycle(graph, neighbor, visited, path):
                return True

        path.pop()
        return False

    def _execute_time_constraint(self, task: SubagentTask) -> Dict[str, Any]:
        """执行时间约束分析"""
        constraint_result = self.results.get("constraint_extraction", {})

        result = {
            "task_id": task.task_id,
            "time_pairs": [],
            "opendate_constraints": [],
            "rules": []
        }

        time_fields = constraint_result.get("constraints", {}).get("time_fields", [])

        # 按表分组
        table_times = {}
        for tf in time_fields:
            table = tf["table"]
            if table not in table_times:
                table_times[table] = []
            table_times[table].append(tf)

        # 检测时间对
        for table, fields in table_times.items():
            field_names = [f["field"].lower() for f in fields]

            if "starttime" in field_names and "endtime" in field_names:
                result["time_pairs"].append({
                    "table": table,
                    "start_field": "StartTime",
                    "end_field": "EndTime",
                    "constraint": "StartTime < EndTime"
                })

            if "opendate" in field_names:
                result["opendate_constraints"].append({
                    "table": table,
                    "field": "OpenDate",
                    "constraint": "IsOpen=true 时 OpenDate 必须有值"
                })

        # 添加通用规则
        result["rules"] = [
            {"rule": "StartTime < EndTime", "description": "开始时间必须早于结束时间"},
            {"rule": "ShowDate <= BeginDate", "description": "显示日期不能晚于开始日期"},
            {"rule": "ExpireDate >= EndDate", "description": "过期日期不能早于结束日期"}
        ]

        return result

    def _execute_data_quality(self, task: SubagentTask) -> Dict[str, Any]:
        """执行数据质量检查"""
        import openpyxl

        scan_result = self.results.get("scan_core", {})

        result = {
            "task_id": task.task_id,
            "issues": {
                "empty_fields": [],
                "duplicate_ids": [],
                "invalid_refs": []
            },
            "stats": {
                "total_rows": 0,
                "total_fields": 0,
                "empty_rate": 0
            }
        }

        total_cells = 0
        empty_cells = 0

        for file_info in scan_result.get("files", []):
            file_path = self.config_dir / file_info["path"]
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                for sheet in wb.worksheets:
                    ids = set()
                    duplicate_ids = set()

                    for row_num, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
                        if not any(cell is not None for cell in row):
                            continue

                        result["stats"]["total_rows"] += 1

                        # 检查ID重复
                        if row[0] is not None:
                            id_val = str(row[0])
                            if id_val in ids:
                                duplicate_ids.add(id_val)
                            ids.add(id_val)

                        # 统计空值
                        for cell in row:
                            total_cells += 1
                            if cell is None:
                                empty_cells += 1

                    if duplicate_ids:
                        result["issues"]["duplicate_ids"].append({
                            "table": sheet.title,
                            "duplicates": list(duplicate_ids)
                        })

                wb.close()
            except:
                pass

        result["stats"]["total_fields"] = total_cells
        if total_cells > 0:
            result["stats"]["empty_rate"] = round(empty_cells / total_cells * 100, 2)

        return result

    def _execute_doc_generation(self, task: SubagentTask) -> Dict[str, Any]:
        """执行文档生成"""
        result = {
            "task_id": task.task_id,
            "output_files": [],
            "summary": {}
        }

        # 收集所有分析结果
        scan_core = self.results.get("scan_core", {})
        scan_enum = self.results.get("scan_enum", {})
        relation = self.results.get("relation_analysis", {})
        constraint = self.results.get("constraint_extraction", {})
        enum_analysis = self.results.get("enum_analysis", {})
        anti_pattern = self.results.get("anti_pattern", {})
        time_constraint = self.results.get("time_constraint", {})
        data_quality = self.results.get("data_quality", {})

        # 生成汇总报告
        result["summary"] = {
            "total_files": scan_core.get("stats", {}).get("total", 0),
            "enum_files": scan_enum.get("stats", {}).get("total", 0),
            "total_rows": scan_core.get("stats", {}).get("rows", 0),
            "relations": len(relation.get("relations", [])),
            "center_tables": len(relation.get("center_tables", [])),
            "orphan_tables": len(anti_pattern.get("orphan_tables", [])),
            "circular_refs": len(anti_pattern.get("circular_refs", [])),
            "time_constraints": len(time_constraint.get("time_pairs", [])),
            "quality_issues": {
                "duplicate_ids": len(data_quality.get("issues", {}).get("duplicate_ids", [])),
                "empty_rate": data_quality.get("stats", {}).get("empty_rate", 0)
            }
        }

        # 生成 Markdown 报告
        md_content = self._generate_markdown_report(result["summary"])

        output_file = self.output_dir / "subagent_analysis_report.md"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(md_content)

        result["output_files"].append(str(output_file))

        return result

    def _generate_markdown_report(self, summary: Dict) -> str:
        """生成 Markdown 报告"""
        return f"""# 配表分析报告（Subagent 调度）

**生成时间**: {time.strftime("%Y-%m-%d %H:%M:%S")}

## 一、统计概览

| 指标 | 数值 |
|------|------|
| 配置表文件数 | {summary.get('total_files', 0)} |
| 枚举表文件数 | {summary.get('enum_files', 0)} |
| 数据总行数 | {summary.get('total_rows', 0)} |
| 表间引用关系 | {summary.get('relations', 0)} |
| 中心表数量 | {summary.get('center_tables', 0)} |

## 二、反模式检测

| 反模式类型 | 数量 |
|------------|------|
| 孤立表 | {summary.get('orphan_tables', 0)} |
| 循环引用 | {summary.get('circular_refs', 0)} |

## 三、约束分析

- 时间约束对: {summary.get('time_constraints', 0)} 组

## 四、数据质量

| 问题类型 | 数量/比例 |
|----------|----------|
| ID重复 | {summary.get('quality_issues', {}).get('duplicate_ids', 0)} 个表 |
| 空值率 | {summary.get('quality_issues', {}).get('empty_rate', 0)}% |

---
*本报告由 Subagent 调度器自动生成*
"""

    def execute_plan(self, plan: AnalysisPlan) -> Dict[str, Any]:
        """执行完整分析计划"""
        print(f"开始执行分析计划: {plan.plan_id}")
        print(f"配置目录: {plan.config_dir}")
        print(f"并行组数: {len(plan.parallel_groups)}")

        start_time = time.time()

        for group_idx, group in enumerate(plan.parallel_groups):
            print(f"\n执行第 {group_idx + 1} 组任务 ({len(group)} 个并行):")
            for task_id in group:
                task = plan.get_task_by_id(task_id)
                if task:
                    print(f"  - {task_id}: {task.description}")

            # 并行执行当前组的任务
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {}
                for task_id in group:
                    task = plan.get_task_by_id(task_id)
                    if task:
                        future = executor.submit(self.execute_task, task)
                        futures[future] = task_id

                for future in as_completed(futures):
                    task_id = futures[future]
                    try:
                        result = future.result()
                        status = "[OK]" if "error" not in result else "[FAIL]"
                        print(f"  {status} {task_id} completed")
                    except Exception as e:
                        print(f"  [FAIL] {task_id} failed: {e}")

        elapsed = time.time() - start_time
        print(f"\n分析完成，耗时: {elapsed:.2f} 秒")

        return {
            "plan_id": plan.plan_id,
            "elapsed_time": elapsed,
            "results": self.results,
            "summary": self.results.get("doc_generation", {}).get("summary", {})
        }


def generate_subagent_prompts(plan: AnalysisPlan) -> Dict[str, str]:
    """为每个任务生成 Claude Agent 提示词"""
    prompts = {}

    for task in plan.tasks:
        prompt = f"""# 配表分析任务: {task.description}

## 任务ID: {task.task_id}

## 任务类型: {task.task_type.value}

## 参数:
```json
{json.dumps(task.params, ensure_ascii=False, indent=2)}
```

## 依赖任务:
{', '.join(task.dependencies) if task.dependencies else '无'}

## 配置目录: {plan.config_dir}

## 输出目录: {plan.output_dir}

## 要求:
1. 分析 D:/work/config/excel 目录下的 Excel 配置文件
2. 配表格式：4行表头（第1行中文、第2行类型、第3行字段名、第4行导出标识）
3. 数据从第5行开始
4. 将结果保存为 JSON 格式到输出目录

## 输出格式:
```json
{{
  "task_id": "{task.task_id}",
  "status": "completed",
  "result": {{}},
  "files_processed": 0
}}
```
"""
        prompts[task.task_id] = prompt

    return prompts


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description="配表分析 Subagent 调度器")
    parser.add_argument("config_dir", help="配置表目录路径")
    parser.add_argument("-o", "--output", help="输出目录路径", default=None)
    parser.add_argument("-w", "--workers", type=int, default=4, help="并行工作线程数")
    parser.add_argument("--generate-prompts", action="store_true", help="生成 Claude Agent 提示词")

    args = parser.parse_args()

    scheduler = SubagentScheduler(
        config_dir=args.config_dir,
        output_dir=args.output,
        max_workers=args.workers
    )

    plan = scheduler.create_full_analysis_plan()

    if args.generate_prompts:
        prompts = generate_subagent_prompts(plan)
        output_dir = Path(args.output) if args.output else Path(args.config_dir)

        for task_id, prompt in prompts.items():
            prompt_file = output_dir / f"prompt_{task_id}.md"
            with open(prompt_file, "w", encoding="utf-8") as f:
                f.write(prompt)
            print(f"生成提示词: {prompt_file}")
    else:
        result = scheduler.execute_plan(plan)
        print(f"\n分析结果: {json.dumps(result['summary'], ensure_ascii=False, indent=2)}")


if __name__ == "__main__":
    main()
