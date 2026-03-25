#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
时间约束提取器

专门用于提取配表中的时间相关约束，特别是 OpenDate 相关约束
"""

import openpyxl
import os
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any
from collections import defaultdict


class TimeConstraintExtractor:
    """时间约束提取器"""

    def __init__(self, config_dir: str):
        self.config_dir = Path(config_dir)
        self.constraints = []
        self.time_fields = defaultdict(list)

    def extract_all_time_constraints(self) -> Dict[str, Any]:
        """提取所有时间约束"""
        result = {
            "constraints": [],
            "tables": {},
            "summary": {
                "total_tables": 0,
                "total_time_fields": 0,
                "constraint_types": defaultdict(int)
            }
        }

        # 遍历所有 Excel 文件
        for file_path in self.config_dir.rglob("*.xlsx"):
            if "~$" in file_path.name:
                continue

            file_constraints = self._extract_file_time_constraints(file_path)
            if file_constraints:
                result["constraints"].extend(file_constraints["constraints"])
                result["tables"][file_path.stem] = file_constraints
                result["summary"]["total_tables"] += 1

        # 统计时间字段总数
        for table_data in result["tables"].values():
            result["summary"]["total_time_fields"] += len(table_data.get("time_fields", []))

        # 统计约束类型
        for constraint in result["constraints"]:
            result["summary"]["constraint_types"][constraint["constraint_type"]] += 1

        return result

    def _extract_file_time_constraints(self, file_path: Path) -> Dict[str, Any]:
        """提取单个文件的时间约束"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            file_result = {
                "filename": file_path.name,
                "path": str(file_path),
                "sheets": {},
                "time_fields": [],
                "constraints": []
            }

            for sheet in wb.worksheets:
                sheet_constraints = self._extract_sheet_time_constraints(sheet, file_path.name)
                if sheet_constraints:
                    file_result["sheets"][sheet.title] = sheet_constraints
                    file_result["time_fields"].extend(sheet_constraints.get("time_fields", []))
                    file_result["constraints"].extend(sheet_constraints.get("constraints", []))

            return file_result if file_result["constraints"] else None

        except Exception as e:
            print(f"警告: 无法读取 {file_path}: {e}")
            return None

    def _extract_sheet_time_constraints(self, sheet, filename: str) -> Dict[str, Any]:
        """提取单个 Sheet 的时间约束"""
        # 找到表头行
        header_row = self._find_header_row(sheet)
        if header_row is None:
            return None

        # 提取表头信息
        headers = self._extract_headers(sheet, header_row)
        if not headers:
            return None

        sheet_result = {
            "sheet_name": sheet.title,
            "time_fields": [],
            "constraints": []
        }

        # 查找时间相关字段
        time_fields = {}
        for header in headers:
            name_lower = header["name"].lower()
            chs_name_lower = header["chs_name"].lower()

            # 识别时间字段
            time_keywords = ['time', 'date', 'start', 'end', 'open', 'close', 'begin', 'finish']
            if any(kw in name_lower or kw in chs_name_lower for kw in time_keywords):
                time_fields[header["name"]] = header
                sheet_result["time_fields"].append(header)

        # 提取各种时间约束
        sheet_result["constraints"].extend(self._extract_time_order_constraints(sheet.title, time_fields))
        sheet_result["constraints"].extend(self._extract_opendate_constraints(sheet.title, time_fields, headers))
        sheet_result["constraints"].extend(self._extract_duration_constraints(sheet.title, time_fields))
        sheet_result["constraints"].extend(self._extract_cross_table_time_constraints(sheet.title, time_fields, filename))

        return sheet_result if sheet_result["constraints"] else None

    def _find_header_row(self, sheet) -> int:
        """查找表头行"""
        # 检查是否有足够的数据
        if sheet.max_row < 4 or sheet.max_column < 1:
            return None

        # 优先查找包含 server/client 的行
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            for cell in sheet[row_idx]:
                if cell.value and 'server/client' in str(cell.value):
                    # 字段名通常在 server/client 行往下第 3 行
                    header_row = row_idx + 3
                    if header_row <= sheet.max_row:
                        return header_row

        # 默认返回第 4 行
        if sheet.max_row >= 4:
            return 4

        return None

    def _extract_headers(self, sheet, header_row: int) -> List[Dict[str, Any]]:
        """提取表头信息"""
        headers = []

        for col_idx in range(1, sheet.max_column + 1):
            chs_name = sheet.cell(1, col_idx).value
            field_type = sheet.cell(2, col_idx).value
            field_name = sheet.cell(header_row, col_idx).value
            export_tag = sheet.cell(header_row + 1, col_idx).value

            if not field_name:
                continue

            headers.append({
                "column": col_idx,
                "chs_name": str(chs_name) if chs_name else "",
                "type": str(field_type) if field_type else "",
                "name": str(field_name),
                "export": str(export_tag) if export_tag else ""
            })

        return headers

    def _extract_time_order_constraints(self, sheet_name: str, time_fields: Dict) -> List[Dict]:
        """提取时间顺序约束"""
        constraints = []

        # StartTime < EndTime
        if "StartTime" in time_fields and "EndTime" in time_fields:
            constraints.append({
                "table": sheet_name,
                "constraint_type": "time_order",
                "rule": "StartTime < EndTime",
                "fields": ["StartTime", "EndTime"],
                "description": "开始时间必须早于结束时间",
                "severity": "error"
            })

        # OpenDate < CloseDate
        if "OpenDate" in time_fields and "CloseDate" in time_fields:
            constraints.append({
                "table": sheet_name,
                "constraint_type": "time_order",
                "rule": "OpenDate < CloseDate",
                "fields": ["OpenDate", "CloseDate"],
                "description": "开启日期必须早于关闭日期",
                "severity": "error"
            })

        return constraints

    def _extract_opendate_constraints(self, sheet_name: str, time_fields: Dict, all_headers: List[Dict]) -> List[Dict]:
        """提取 OpenDate 相关约束"""
        constraints = []

        if "OpenDate" not in time_fields:
            return constraints

        open_date_field = time_fields["OpenDate"]

        # 约束1: IsOpen=true 时 OpenDate 必须有值
        is_open_exists = any(h["name"] == "IsOpen" for h in all_headers)
        if is_open_exists:
            constraints.append({
                "table": sheet_name,
                "constraint_type": "opendate_conditional",
                "rule": "IsOpen=true 时 OpenDate 必须有值",
                "fields": ["OpenDate", "IsOpen"],
                "description": "当开启标志为 true 时，必须指定开启日期",
                "severity": "error",
                "condition": "IsOpen == true"
            })

        # 约束2: OpenDate 格式约束（通常是 YYYY-MM-DD HH:mm:ss）
        constraints.append({
            "table": sheet_name,
            "constraint_type": "opendate_format",
            "rule": "OpenDate 格式必须为 YYYY-MM-DD HH:mm:ss 或时间戳",
            "fields": ["OpenDate"],
            "description": "开启日期必须使用标准时间格式",
            "severity": "warning",
            "valid_formats": ["YYYY-MM-DD HH:mm:ss", "Unix timestamp"]
        })

        # 约束3: OpenDate 不能早于服务器启动时间
        constraints.append({
            "table": sheet_name,
            "constraint_type": "opendate_future",
            "rule": "OpenDate >= 服务器首次启动时间",
            "fields": ["OpenDate"],
            "description": "开启日期不能早于服务器首次启动时间",
            "severity": "warning",
            "reference": "服务器启动时间"
        })

        # 约束4: 大将军武将的 OpenDate 约束（如果是 Hero 表）
        if "Hero" in sheet_name or "hero" in sheet_name.lower():
            constraints.append({
                "table": sheet_name,
                "constraint_type": "opendate_hero_general",
                "rule": "大将军武将的 OpenDate 必须与战令赛季同步",
                "fields": ["OpenDate", "Quality", "GeneralType"],
                "description": "大将军武将的开启时间应该与对应战令赛季的开始时间一致",
                "severity": "warning",
                "condition": "Quality == '大将军' 或 GeneralType indicates 大将军",
                "reference_table": "SeasonPass",
                "reference_field": "StartTime"
            })

            # 约束5: 战令武将识别规则
            constraints.append({
                "table": sheet_name,
                "constraint_type": "opendate_season_pass_hero",
                "rule": "战令武将 OpenDate 与 SeasonPass 表的 StartTime 一致",
                "fields": ["OpenDate"],
                "description": "战令武将的 OpenDate 应该对应到 SeasonPass 表中的 StartTime",
                "severity": "info",
                "detection_rule": "通过特定标识识别战令武将（如特定前缀、标签或枚举值）",
                "reference_table": "SeasonPass",
                "reference_field": "StartTime"
            })

        return constraints

    def _extract_duration_constraints(self, sheet_name: str, time_fields: Dict) -> List[Dict]:
        """提取持续时间约束"""
        constraints = []

        # 持续时间字段约束
        if "Duration" in time_fields or "ContinueTime" in time_fields:
            duration_field = time_fields.get("Duration") or time_fields.get("ContinueTime")
            constraints.append({
                "table": sheet_name,
                "constraint_type": "duration_positive",
                "rule": "Duration > 0",
                "fields": [duration_field["name"]],
                "description": "持续时间必须大于0",
                "severity": "error"
            })

        return constraints

    def _extract_cross_table_time_constraints(self, sheet_name: str, time_fields: Dict, filename: str) -> List[Dict]:
        """提取跨表时间约束"""
        constraints = []

        # Hero 表的 OpenDate 与 SeasonPass 的关系
        if "OpenDate" in time_fields and ("Hero" in sheet_name or "hero" in sheet_name.lower()):
            constraints.append({
                "table": sheet_name,
                "constraint_type": "cross_table_time_sync",
                "rule": "Hero.OpenDate ∈ SeasonPass 时间范围",
                "fields": ["OpenDate"],
                "description": "武将开启时间应该在某个战令赛季的时间范围内",
                "severity": "warning",
                "reference_table": "SeasonPass",
                "reference_fields": ["StartTime", "EndTime"],
                "relationship": "Hero.OpenDate >= SeasonPass.StartTime AND Hero.OpenDate <= SeasonPass.EndTime"
            })

        # ArenaSeason 相关的时间约束
        if "ArenaSeason" in sheet_name or "arena" in sheet_name.lower():
            if "StartTime" in time_fields and "EndTime" in time_fields:
                constraints.append({
                    "table": sheet_name,
                    "constraint_type": "season_time_span",
                    "rule": "赛季持续时间通常为 7-30 天",
                    "fields": ["StartTime", "EndTime"],
                    "description": "竞技场赛季的持续时间建议在 7-30 天之间",
                    "severity": "info",
                    "suggested_duration": "7-30 天"
                })

        return constraints

    def generate_constraint_report(self, result: Dict[str, Any]) -> str:
        """生成约束报告"""
        lines = []
        lines.append("# 时间约束分析报告")
        lines.append("")
        lines.append(f"## 总览")
        lines.append("")
        lines.append(f"- 分析表数量: {result['summary']['total_tables']}")
        lines.append(f"- 时间字段总数: {result['summary']['total_time_fields']}")
        lines.append(f"- 约束规则总数: {len(result['constraints'])}")
        lines.append("")
        lines.append("## 约束类型统计")
        lines.append("")
        lines.append("| 约束类型 | 数量 |")
        lines.append("|----------|------|")

        for constraint_type, count in result['summary']['constraint_types'].items():
            lines.append(f"| {constraint_type} | {count} |")

        lines.append("")
        lines.append("## 详细约束列表")
        lines.append("")

        # 按表分组
        table_groups = defaultdict(list)
        for constraint in result['constraints']:
            table_groups[constraint['table']].append(constraint)

        for table_name, constraints in sorted(table_groups.items()):
            lines.append(f"### {table_name}")
            lines.append("")

            for idx, constraint in enumerate(constraints, 1):
                lines.append(f"#### {idx}. {constraint['constraint_type']}")
                lines.append("")
                lines.append(f"- **规则**: {constraint['rule']}")
                lines.append(f"- **涉及字段**: {', '.join(constraint['fields'])}")
                lines.append(f"- **描述**: {constraint['description']}")
                lines.append(f"- **严重度**: {constraint['severity']}")

                if 'condition' in constraint:
                    lines.append(f"- **条件**: {constraint['condition']}")

                if 'reference_table' in constraint:
                    lines.append(f"- **关联表**: {constraint['reference_table']}")
                    if 'reference_field' in constraint:
                        lines.append(f"- **关联字段**: {constraint['reference_field']}")
                    if 'reference_fields' in constraint:
                        lines.append(f"- **关联字段**: {', '.join(constraint['reference_fields'])}")

                if 'valid_formats' in constraint:
                    lines.append(f"- **有效格式**: {', '.join(constraint['valid_formats'])}")

                if 'suggested_duration' in constraint:
                    lines.append(f"- **建议值**: {constraint['suggested_duration']}")

                lines.append("")

        return "\n".join(lines)

    def generate_mermaid_constraint_graph(self, result: Dict[str, Any]) -> str:
        """生成约束关系图"""
        lines = []
        lines.append("graph TD")

        # 收集所有表和约束
        tables = set()
        for constraint in result['constraints']:
            tables.add(constraint['table'])

        # 生成节点
        for table in sorted(tables):
            lines.append(f'    {table}[{table}]')

        lines.append("")

        # 生成关系（基于跨表约束）
        cross_table_constraints = [c for c in result['constraints'] if c['constraint_type'] == 'cross_table_time_sync']
        for constraint in cross_table_constraints:
            source = constraint['table']
            target = constraint.get('reference_table', 'Unknown')
            lines.append(f'    {source} -->|{constraint["rule"]}| {target}')

        return "\n".join(lines)


def main():
    """主函数"""
    import sys

    if len(sys.argv) < 3:
        print("用法: python extract_time_constraints.py <配表目录> <输出目录>")
        sys.exit(1)

    config_dir = sys.argv[1]
    output_dir = sys.argv[2]

    extractor = TimeConstraintExtractor(config_dir)

    print("提取时间约束...")
    result = extractor.extract_all_time_constraints()

    print(f"找到 {result['summary']['total_tables']} 个表")
    print(f"找到 {result['summary']['total_time_fields']} 个时间字段")
    print(f"找到 {len(result['constraints'])} 个约束规则")

    # 生成报告
    print("生成报告...")
    report = extractor.generate_constraint_report(result)

    # 生成图表
    print("生成图表...")
    graph = extractor.generate_mermaid_constraint_graph(result)

    # 保存结果
    os.makedirs(output_dir, exist_ok=True)

    # 保存 JSON 结果
    json_output = os.path.join(output_dir, "time_constraints.json")
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    # 保存 Markdown 报告
    report_output = os.path.join(output_dir, "time_constraints_report.md")
    with open(report_output, 'w', encoding='utf-8') as f:
        f.write(report)

    # 保存 Mermaid 图表
    graph_output = os.path.join(output_dir, "time_constraints_graph.mmd")
    with open(graph_output, 'w', encoding='utf-8') as f:
        f.write(graph)

    # 保存 Hero 表专门的约束
    hero_constraints = [c for c in result['constraints'] if 'Hero' in c['table'] or 'hero' in c['table'].lower()]
    if hero_constraints:
        hero_output = os.path.join(output_dir, "hero_time_constraints.json")
        with open(hero_output, 'w', encoding='utf-8') as f:
            json.dump(hero_constraints, f, ensure_ascii=False, indent=2)

    print(f"\n结果已保存到: {output_dir}")
    print(f"- JSON: {json_output}")
    print(f"- 报告: {report_output}")
    print(f"- 图表: {graph_output}")
    if hero_constraints:
        print(f"- Hero 约束: {hero_output}")


if __name__ == "__main__":
    main()
