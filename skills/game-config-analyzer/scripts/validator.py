#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配表验证引擎

执行数据完整性、一致性和业务规则验证
"""

import openpyxl
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Any, Optional
from collections import defaultdict


class ConfigValidator:
    """配表验证器"""

    def __init__(self, config_dir: str):
        self.config_dir = Path(config_dir)
        self.tables = {}
        self.relations = []
        self.constraints = []

    def load_analysis(self, analysis_file: str):
        """加载分析结果"""
        with open(analysis_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            self.relations = data.get("relations", [])
            self.constraints = data.get("constraints", [])

    def validate_all(self) -> Dict[str, Any]:
        """执行所有验证"""
        print("扫描配表...")
        self._scan_tables()

        print("验证数据完整性...")
        integrity_errors = self.validate_integrity()

        print("验证数据一致性...")
        consistency_errors = self.validate_consistency()

        print("验证业务规则...")
        business_errors = self.validate_business_rules()

        # 汇总结果
        all_errors = integrity_errors + consistency_errors + business_errors

        result = {
            "timestamp": datetime.now().isoformat(),
            "summary": {
                "total_errors": len(all_errors),
                "by_severity": self._count_by_severity(all_errors),
                "by_type": self._count_by_type(all_errors),
                "by_table": self._count_by_table(all_errors)
            },
            "errors": all_errors
        }

        return result

    def _scan_tables(self):
        """扫描配表"""
        for file_path in self.config_dir.rglob("*.xlsx"):
            if "~$" in file_path.name:
                continue

            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet in wb.worksheets:
                    table_data = self._extract_table_data(sheet)
                    if table_data:
                        key = f"{file_path.name}/{sheet.title}"
                        self.tables[key] = table_data
            except Exception as e:
                print(f"警告: 无法读取 {file_path}: {e}")

    def _extract_table_data(self, sheet) -> Optional[Dict]:
        """提取表数据"""
        # 假设标准 4 行表头
        header_row = 3
        data_start_row = 4

        # 提取表头
        headers = []
        for col in range(1, sheet.max_column + 1):
            field_name = sheet.cell(header_row, col).value
            if field_name:
                headers.append({
                    "column": col,
                    "name": str(field_name),
                    "type": str(sheet.cell(2, col).value) if sheet.cell(2, col).value else ""
                })

        if not headers:
            return None

        # 提取数据
        data = []
        id_field = None
        id_index = self._find_id_index(headers)

        for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not any(cell is not None for cell in row):
                continue

            row_data = {}
            for col_idx, header in enumerate(headers):
                col = header["column"] - 1
                value = row[col] if col < len(row) else None
                row_data[header["name"]] = value

            row_data["_row"] = row_idx
            data.append(row_data)

        return {
            "headers": headers,
            "data": data,
            "id_field": headers[id_index]["name"] if id_index is not None else None
        }

    def _find_id_index(self, headers: List[Dict]) -> Optional[int]:
        """查找 ID 字段索引"""
        for idx, header in enumerate(headers):
            if header["name"] == "Id":
                return idx
        return None

    def validate_integrity(self) -> List[Dict]:
        """验证数据完整性"""
        errors = []

        # 构建 ID 索引
        id_index = self._build_id_index()

        # 验证外键引用
        for rel in self.relations:
            errors.extend(self._validate_reference(rel, id_index))

        # 验证必填字段
        for table_key, table_data in self.tables.items():
            errors.extend(self._validate_required_fields(table_key, table_data))

        # 验证唯一性
        for table_key, table_data in self.tables.items():
            errors.extend(self._validate_uniqueness(table_key, table_data))

        return errors

    def _build_id_index(self) -> Dict[str, Set]:
        """构建 ID 索引"""
        index = defaultdict(set)

        for table_key, table_data in self.tables.items():
            if table_data["id_field"]:
                field_name = table_data["id_field"]
                table_name = table_key.split('/')[0].replace('.xlsx', '')

                for row in table_data["data"]:
                    id_value = row.get(field_name)
                    if id_value is not None:
                        index[f"{table_name}.{field_name}"].add(id_value)

        return index

    def _validate_reference(self, rel: Dict, id_index: Dict) -> List[Dict]:
        """验证引用完整性"""
        errors = []
        source_table = rel["source_table"]
        source_field = rel["source_field"]
        target_table = rel["target_table"]
        target_field = rel["target_field"]

        # 找到源表数据
        source_table_key = self._find_table_key(source_table)
        if not source_table_key:
            return errors

        table_data = self.tables[source_table_key]
        target_key = f"{target_table}.{target_field}"
        valid_ids = id_index.get(target_key, set())

        for row in table_data["data"]:
            ref_value = row.get(source_field)
            if ref_value is not None and ref_value not in valid_ids:
                errors.append({
                    "severity": "error",
                    "type": "broken_reference",
                    "table": source_table,
                    "row": row["_row"],
                    "field": source_field,
                    "value": ref_value,
                    "target": f"{target_table}.{target_field}",
                    "message": f"引用的 ID {ref_value} 在 {target_table} 中不存在"
                })

        return errors

    def _validate_required_fields(self, table_key: str, table_data: Dict) -> List[Dict]:
        """验证必填字段"""
        errors = []
        table_name = table_key.split('/')[0]

        # 识别必填字段（导出标记为 server 的字段）
        required_fields = []
        for header in table_data["headers"]:
            # 这里简化处理，实际需要检查第 4 行的导出标记
            field_name = header["name"]
            if "Id" in field_name or "Name" in field_name:
                required_fields.append(field_name)

        for row in table_data["data"]:
            for field in required_fields:
                if row.get(field) is None:
                    errors.append({
                        "severity": "error",
                        "type": "missing_required",
                        "table": table_name,
                        "row": row["_row"],
                        "field": field,
                        "message": f"必填字段 {field} 为空"
                    })

        return errors

    def _validate_uniqueness(self, table_key: str, table_data: Dict) -> List[Dict]:
        """验证唯一性"""
        errors = []
        table_name = table_key.split('/')[0]

        if not table_data["id_field"]:
            return errors

        id_field = table_data["id_field"]
        seen_ids = {}

        for row in table_data["data"]:
            id_value = row.get(id_field)
            if id_value is not None:
                if id_value in seen_ids:
                    errors.append({
                        "severity": "error",
                        "type": "duplicate_key",
                        "table": table_name,
                        "field": id_field,
                        "value": id_value,
                        "rows": [seen_ids[id_value], row["_row"]],
                        "message": f"主键 {id_value} 重复出现在行 {seen_ids[id_value]} 和 {row['_row']}"
                    })
                else:
                    seen_ids[id_value] = row["_row"]

        return errors

    def validate_consistency(self) -> List[Dict]:
        """验证数据一致性"""
        errors = []

        for table_key, table_data in self.tables.items():
            # 时间顺序验证
            errors.extend(self._validate_time_order(table_key, table_data))

        return errors

    def _validate_time_order(self, table_key: str, table_data: Dict) -> List[Dict]:
        """验证时间顺序"""
        errors = []
        table_name = table_key.split('/')[0]

        # 查找时间字段
        headers = {h["name"]: h for h in table_data["headers"]}
        time_fields = [name for name in headers.keys()
                      if any(kw in name.lower() for kw in ["starttime", "endtime", "time", "date"])]

        has_start = any("starttime" in f.lower() for f in time_fields)
        has_end = any("endtime" in f.lower() for f in time_fields)

        if has_start and has_end:
            # 找到具体字段名
            start_field = next((f for f in time_fields if "starttime" in f.lower()), None)
            end_field = next((f for f in time_fields if "endtime" in f.lower()), None)

            if start_field and end_field:
                for row in table_data["data"]:
                    start_val = row.get(start_field)
                    end_val = row.get(end_field)

                    if start_val and end_val:
                        try:
                            if isinstance(start_val, str) and isinstance(end_val, str):
                                # 简单的字符串比较
                                if start_val >= end_val:
                                    errors.append({
                                        "severity": "error",
                                        "type": "time_order",
                                        "table": table_name,
                                        "row": row["_row"],
                                        "message": f"{start_field} ({start_val}) 必须 < {end_field} ({end_val})"
                                    })
                        except:
                            pass

        return errors

    def validate_business_rules(self) -> List[Dict]:
        """验证业务规则"""
        errors = []

        # 从预定义的约束中验证
        for constraint in self.constraints:
            errors.extend(self._validate_constraint(constraint))

        return errors

    def _validate_constraint(self, constraint: Dict) -> List[Dict]:
        """验证单个约束"""
        errors = []
        table_name = constraint["table"]
        constraint_type = constraint.get("constraint_type", "")

        table_key = self._find_table_key(table_name)
        if not table_key:
            return errors

        table_data = self.tables[table_key]

        if constraint_type == "time_order":
            # StartTime < EndTime
            fields = constraint.get("fields", [])
            if len(fields) >= 2:
                for row in table_data["data"]:
                    val1 = row.get(fields[0])
                    val2 = row.get(fields[1])
                    if val1 and val2 and val1 >= val2:
                        errors.append({
                            "severity": "error",
                            "type": "constraint_violation",
                            "table": table_name,
                            "row": row["_row"],
                            "constraint": constraint["rule"],
                            "message": f"违反约束: {constraint['rule']}"
                        })

        return errors

    def _find_table_key(self, table_name: str) -> Optional[str]:
        """查找表键"""
        for key in self.tables.keys():
            if table_name in key:
                return key
        return None

    def _count_by_severity(self, errors: List[Dict]) -> Dict[str, int]:
        """按严重程度统计"""
        counts = defaultdict(int)
        for error in errors:
            counts[error.get("severity", "unknown")] += 1
        return dict(counts)

    def _count_by_type(self, errors: List[Dict]) -> Dict[str, int]:
        """按类型统计"""
        counts = defaultdict(int)
        for error in errors:
            counts[error.get("type", "unknown")] += 1
        return dict(counts)

    def _count_by_table(self, errors: List[Dict]) -> Dict[str, int]:
        """按表统计"""
        counts = defaultdict(int)
        for error in errors:
            table = error.get("table", "unknown")
            counts[table] += 1
        return dict(counts)

    def generate_report(self, result: Dict) -> str:
        """生成验证报告"""
        lines = [
            "# 配表验证报告",
            f"",
            f"**生成时间**: {result['timestamp']}",
            f"",
            f"## 验证摘要",
            f"",
            f"- 总问题数: {result['summary']['total_errors']}",
            f""
        ]

        # 按严重程度
        by_severity = result['summary']['by_severity']
        if by_severity:
            lines.extend([
                f"### 按严重程度",
                f"",
                f"| 严重程度 | 数量 |",
                f"|----------|------|"
            ])
            for severity, count in sorted(by_severity.items(), key=lambda x: -len(x[0])):
                icon = {"error": "🔴", "warning": "🟡", "info": "🟢"}.get(severity, "")
                lines.append(f"| {icon} {severity.upper()} | {count} |")
            lines.append("")

        # 按类型
        by_type = result['summary']['by_type']
        if by_type:
            lines.extend([
                f"### 按问题类型",
                f"",
                f"| 类型 | 数量 |",
                f"|------|------|"
            ])
            for error_type, count in sorted(by_type.items(), key=lambda x: -x[1]):
                lines.append(f"| {error_type} | {count} |")
            lines.append("")

        # 按表
        by_table = result['summary']['by_table']
        if by_table:
            lines.extend([
                f"### 按表",
                f"",
                f"| 表 | 问题数 |",
                f"|----|-------|"
            ])
            for table, count in sorted(by_table.items(), key=lambda x: -x[1])[:20]:
                lines.append(f"| {table} | {count} |")
            lines.append("")

        # 详细错误
        if result["errors"]:
            lines.extend([
                f"## 详细错误",
                f""
            ])
            for error in result["errors"][:100]:  # 只显示前 100 个
                severity = error.get("severity", "unknown")
                icon = {"error": "🔴", "warning": "🟡", "info": "🟢"}.get(severity, "⚪")
                lines.append(f"### {icon} [{severity.upper()}] {error.get('table', 'Unknown')}")
                lines.append(f"**类型**: {error.get('type', 'Unknown')}")
                lines.append(f"**信息**: {error.get('message', 'No message')}")

                if "row" in error:
                    lines.append(f"**行**: {error['row']}")
                if "field" in error:
                    lines.append(f"**字段**: {error['field']}")
                if "value" in error:
                    lines.append(f"**值**: {error['value']}")

                lines.append("")

            if len(result["errors"]) > 100:
                lines.append(f"\n... 还有 {len(result['errors']) - 100} 个错误未显示\n")

        return "\n".join(lines)


def main():
    """主函数"""
    import sys

    if len(sys.argv) < 2:
        print("用法: python validator.py <配表目录> [分析文件.json] [输出文件]")
        sys.exit(1)

    config_dir = sys.argv[1]
    analysis_file = sys.argv[2] if len(sys.argv) > 2 else None
    output_file = sys.argv[3] if len(sys.argv) > 3 else "validation_result.json"

    validator = ConfigValidator(config_dir)

    # 加载分析结果（如果有）
    if analysis_file:
        validator.load_analysis(analysis_file)

    # 执行验证
    result = validator.validate_all()

    # 保存结果
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"JSON 结果已保存到: {output_file}")

    # 生成报告
    report = validator.generate_report(result)
    report_file = output_file.replace('.json', '_report.md')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"Markdown 报告已保存到: {report_file}")

    # 打印摘要
    print()
    print("验证摘要:")
    print(f"  总问题数: {result['summary']['total_errors']}")
    for severity, count in result['summary']['by_severity'].items():
        print(f"  {severity.upper()}: {count}")


if __name__ == "__main__":
    main()
