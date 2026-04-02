#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配表分析辅助脚本

用于扫描配表文件、分析关系、生成报告
"""

import openpyxl
import os
import json
import re
import random
import yaml
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any, Optional
from collections import defaultdict, Counter


class ConfigAnalyzer:
    """配表分析器"""

    def __init__(self, config_dir: str, project_root: str = None):
        self.config_dir = Path(config_dir)
        self.project_root = Path(project_root) if project_root else self.config_dir.parent
        self.tables = {}  # {filename: {sheets, structure}}
        self.relations = []  # 关系列表
        self.constraints = []  # 约束列表
        self.project_config = self.load_project_config()

    def scan_directory(self) -> Dict[str, Any]:
        """扫描目录，收集所有配表文件"""
        result = {
            "files": [],
            "total_count": 0,
            "sheets_count": 0
        }

        for file_path in self.config_dir.rglob("*.xlsx"):
            # 跳过临时文件
            if "~$" in file_path.name:
                continue

            file_info = self._analyze_file(file_path)
            if file_info:
                result["files"].append(file_info)
                result["total_count"] += 1
                result["sheets_count"] += len(file_info["sheets"])

        return result

    def _analyze_file(self, file_path: Path) -> Dict[str, Any]:
        """分析单个文件"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            sheets = []

            for sheet in wb.worksheets:
                sheet_info = self._analyze_sheet(sheet)
                if sheet_info:
                    sheets.append(sheet_info)

            if not sheets:
                return None

            return {
                "filename": file_path.name,
                "path": str(file_path),
                "sheets": sheets
            }
        except Exception as e:
            print(f"警告: 无法读取 {file_path}: {e}")
            return None

    def _analyze_sheet(self, sheet) -> Dict[str, Any]:
        """分析单个 Sheet"""
        # 找到数据起始行
        header_row = self._find_header_row(sheet)
        if header_row is None:
            return None

        # 提取表头信息
        headers = self._extract_headers(sheet, header_row)
        if not headers:
            return None

        # 统计数据行数
        data_start_row = header_row + 1
        data_rows = 0
        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            if row and any(cell is not None for cell in row):
                data_rows += 1

        return {
            "name": sheet.title,
            "headers": headers,
            "row_count": data_rows,
            "header_row": header_row
        }

    def _find_header_row(self, sheet) -> int:
        """查找表头行（包含字段名的那一行）

        优先使用项目配置，回退到自动探测，最后使用默认假设。
        """
        # 优先读取项目配置
        if self.project_config:
            header_conf = self.project_config.get("header", {})
            field_row = header_conf.get("field_name_row")
            if field_row is not None:
                return field_row

        # 回退到自动探测（原有逻辑）
        return self._find_header_row_legacy(sheet)

    def _find_header_row_legacy(self, sheet) -> int:
        """旧版表头检测逻辑，作为无配置时的回退"""
        # 检测标准4行表头：第3行应包含英文字段名，第4行应包含 server/client
        if sheet.max_row < 4:
            return None

        # 检查第4行是否有 server/client 标识（标准格式的特征）
        has_export_tag = False
        for col in range(1, min(sheet.max_column + 1, 20)):
            val = sheet.cell(4, col).value
            if val and ('server' in str(val).lower() or 'client' in str(val).lower()):
                has_export_tag = True
                break

        if has_export_tag:
            return 3  # 字段名在第3行

        # 尝试检测第3行是否为英文字段名（字母开头，不含中文）
        has_field_names = False
        for col in range(1, min(sheet.max_column + 1, 20)):
            val = sheet.cell(3, col).value
            if val and re.match(r'^[A-Za-z_]', str(val)):
                has_field_names = True
                break

        if has_field_names:
            return 3

        return None  # 无法识别表头格式

    def _extract_headers(self, sheet, header_row: int) -> List[Dict[str, Any]]:
        """提取表头信息

        标准4行表头格式（header_row=3，即字段名在第3行）：
        第1行 = 中文名称
        第2行 = 字段类型（int, string, bool, E#枚举名, #注释）
        第3行 = 字段名（英文字段名，如 Id, Name, SkillId）
        第4行 = 导出标识（server/client/server/client）
        """
        headers = []

        for col_idx in range(1, sheet.max_column + 1):
            chs_name = sheet.cell(1, col_idx).value       # 第1行：中文名
            field_type = sheet.cell(2, col_idx).value      # 第2行：类型
            field_name = sheet.cell(header_row, col_idx).value   # 第3行：字段名
            export_tag = sheet.cell(header_row + 1, col_idx).value  # 第4行：导出标识

            if not field_name:
                continue

            # 跳过注释列（类型为 #）
            if field_type and str(field_type).strip() == '#':
                continue

            headers.append({
                "column": col_idx,
                "chs_name": str(chs_name) if chs_name else "",
                "type": str(field_type).strip() if field_type else "",
                "name": str(field_name).strip(),
                "export": str(export_tag).strip() if export_tag else ""
            })

        return headers

    def analyze_relations(self, scan_result: Dict[str, Any]) -> List[Dict[str, Any]]:
        """分析表之间的引用关系"""
        relations = []

        # 构建表名到文件的映射
        # 支持两种 Sheet 名格式：纯英文名（Hero）和中英混合（成就表|Achieve）
        table_map = {}
        for file_info in scan_result["files"]:
            for sheet_info in file_info["sheets"]:
                sheet_name = sheet_info["name"]

                # 提取英文名部分（| 后面的部分）
                if '|' in sheet_name:
                    eng_name = sheet_name.split('|')[-1].strip()
                else:
                    eng_name = sheet_name.replace('.xlsx', '')

                # 同时用文件名（不含扩展名）和Sheet英文名建立索引
                file_base = file_info["filename"].replace('.xlsx', '')
                for name in [eng_name, file_base]:
                    if name and name not in table_map:
                        table_map[name] = {
                            "file": file_info["filename"],
                            "sheet": sheet_info["name"],
                            "headers": sheet_info["headers"]
                        }

        # 分析 ID 引用关系
        for file_info in scan_result["files"]:
            for sheet_info in file_info["sheets"]:
                for header in sheet_info["headers"]:
                    relations.extend(self._find_id_references(
                        header, sheet_info, table_map
                    ))

        return relations

    def _find_id_references(self, header: Dict, sheet_info: Dict, table_map: Dict) -> List[Dict]:
        """查找 ID 引用关系"""
        relations = []
        field_name = header["name"]

        # 模式1: XxxId -> Xxx.Id
        id_match = re.match(r'(.+)Id$', field_name)
        if id_match:
            ref_table = id_match.group(1)
            if ref_table in table_map:
                relations.append({
                    "source_table": sheet_info["name"],
                    "source_field": field_name,
                    "target_table": ref_table,
                    "target_field": "Id",
                    "type": "id_reference"
                })

        # 模式2: XxxIds -> Xxx.Id (数组引用)
        ids_match = re.match(r'(.+)Ids$', field_name)
        if ids_match:
            ref_table = ids_match.group(1)
            if ref_table in table_map:
                relations.append({
                    "source_table": sheet_info["name"],
                    "source_field": field_name,
                    "target_table": ref_table,
                    "target_field": "Id",
                    "type": "id_array_reference"
                })

        return relations

    def extract_constraints(self, scan_result: Dict[str, Any]) -> List[Dict[str, Any]]:
        """提取约束规则"""
        constraints = []

        for file_info in scan_result["files"]:
            for sheet_info in file_info["sheets"]:
                constraints.extend(self._extract_sheet_constraints(sheet_info))

        return constraints

    def _extract_sheet_constraints(self, sheet_info: Dict) -> List[Dict]:
        """提取 Sheet 级别的约束"""
        constraints = []
        headers = sheet_info["headers"]
        sheet_name = sheet_info["name"]

        # 查找时间相关字段
        time_fields = {}
        for header in headers:
            name = header["name"].lower()
            if any(kw in name for kw in ['time', 'date', 'start', 'end']):
                time_fields[header["name"]] = header

        # 时间顺序约束
        if "StartTime" in time_fields and "EndTime" in time_fields:
            constraints.append({
                "table": sheet_name,
                "constraint_type": "time_order",
                "rule": "StartTime < EndTime",
                "fields": ["StartTime", "EndTime"]
            })

        # 时间范围约束
        if "OpenDate" in time_fields:
            constraints.append({
                "table": sheet_name,
                "constraint_type": "opendate_requirement",
                "rule": "IsOpen=true 时 OpenDate 必须有值",
                "fields": ["OpenDate", "IsOpen"]
            })

        return constraints

    def generate_mermaid_graph(self, relations: List[Dict]) -> str:
        """生成 Mermaid 关系图"""
        lines = ["graph TD"]

        # 收集所有节点
        nodes = set()
        for rel in relations:
            nodes.add(rel["source_table"])
            nodes.add(rel["target_table"])

        # 生成节点
        for node in sorted(nodes):
            lines.append(f'    {node}[{node}]')

        # 生成关系
        for rel in relations:
            style = ""
            if rel["type"] == "id_array_reference":
                style = " -.->|数组引用|"
            else:
                style = " -->|引用|"
            lines.append(f'    {rel["source_table"]}{style}{rel["target_table"]}')

        return "\n".join(lines)

    def generate_test_data_template(self, scan_result: Dict) -> str:
        """生成测试数据模板"""
        template = """# 测试用例模板

## 测试场景矩阵

| 场景ID | 表名 | 字段 | 测试值 | 期望结果 |
|--------|------|------|--------|----------|
"""

        return template


    # ========== 自适应初始化方法 ==========

    def load_project_config(self) -> Optional[Dict]:
        """从项目根目录加载 .gameconfig.yaml 配置文件"""
        config_path = self.project_root / ".gameconfig.yaml"
        if not config_path.exists():
            return None
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except Exception as e:
            print(f"警告: 无法读取配置文件 {config_path}: {e}")
            return None

    def save_project_config(self, config: Dict, path: Path = None):
        """保存项目配置到 .gameconfig.yaml"""
        target = path or (self.project_root / ".gameconfig.yaml")
        with open(target, 'w', encoding='utf-8') as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
        print(f"配置已保存到: {target}")

    def _default_config(self) -> Dict:
        """返回默认假设配置（4行表头格式）"""
        return {
            "version": "1.0",
            "project": {"name": self.project_root.name, "detected_at": datetime.now().isoformat(), "detector_version": "1.0"},
            "header": {
                "total_rows": 4, "data_start_row": 5, "field_name_row": 3,
                "rows": {1: "description", 2: "type", 3: "field_name", 4: "export_tag"},
                "confidence": 0.0, "samples_used": 0, "overrides": {}
            },
            "references": {
                "naming_patterns": [
                    {"pattern": "^(.+)Id$", "target": "{1}.Id", "type": "single_reference"},
                    {"pattern": "^(.+)Ids$", "target": "{1}.Id", "type": "array_reference"}
                ],
                "type_patterns": [
                    {"match": "^E[A-Z]", "reference_type": "enum"},
                    {"match": "^#", "behavior": "comment_column"},
                    {"match": "\\[\\]$", "reference_type": "array"}
                ],
                "discovered_patterns": []
            },
            "sheet_naming": {"separator": "|", "use_english_part": True},
            "directories": {"excel": str(self.config_dir.name), "enum": None, "output": "docs"}
        }

    def auto_detect_header_format(self, sample_count: int = 10) -> Dict:
        """自适应检测项目配表的表头格式

        算法：采样 → 逐行角色评分 → 跨文件投票 → 输出配置
        """
        # 收集所有 xlsx 文件
        all_files = [f for f in self.config_dir.rglob("*.xlsx") if "~$" not in f.name]
        if not all_files:
            print("未找到 xlsx 文件，使用默认配置")
            return self._default_config()

        # 采样：选不同大小的文件确保代表性
        sampled = self._representative_sample(all_files, sample_count)
        print(f"采样 {len(sampled)}/{len(all_files)} 个文件进行格式探测...")

        # 逐文件分析行角色
        per_file_results = []
        for file_path in sampled:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False)
                for sheet in wb.worksheets:
                    roles = self._detect_row_roles(sheet, max_scan_rows=8)
                    per_file_results.append({"file": file_path.name, "sheet": sheet.title, "roles": roles})
                    break  # 只取第一个 sheet
                wb.close()
            except Exception as e:
                print(f"  跳过 {file_path.name}: {e}")

        if not per_file_results:
            return self._default_config()

        # 跨文件投票确定行角色
        header_config = self._vote_on_roles(per_file_results)
        print(f"检测到表头格式: {header_config['total_rows']} 行表头，字段名在第 {header_config['field_name_row']} 行")

        # 引用模式发现
        scan_result = self.scan_directory()
        ref_patterns = self._detect_reference_patterns(scan_result)
        print(f"发现 {len(ref_patterns['naming_patterns'])} 种引用模式")

        # Sheet 命名约定检测
        sheet_naming = self._detect_sheet_naming(scan_result)

        # 异常文件覆盖
        overrides = self._detect_overrides(per_file_results, header_config)

        # 置信度计算
        confidence = self._calculate_confidence(header_config, per_file_results)
        print(f"检测置信度: {confidence:.0%}")

        return {
            "version": "1.0",
            "project": {"name": self.project_root.name, "detected_at": datetime.now().isoformat(), "detector_version": "1.0"},
            "header": {**header_config, "confidence": confidence, "samples_used": len(sampled), "overrides": overrides},
            "references": ref_patterns,
            "sheet_naming": sheet_naming,
            "directories": {"excel": str(self.config_dir.name), "enum": None, "output": "docs"}
        }

    def _representative_sample(self, files: List[Path], count: int) -> List[Path]:
        """从文件列表中选取代表性样本（按文件大小分层）"""
        sorted_files = sorted(files, key=lambda f: f.stat().st_size)
        if len(sorted_files) <= count:
            return sorted_files
        # 均匀采样：从不同大小层各取一个
        step = len(sorted_files) / count
        return [sorted_files[int(i * step)] for i in range(count)]

    def _detect_row_roles(self, sheet, max_scan_rows: int = 8) -> Dict:
        """分析单个 sheet 前 N 行的角色"""
        results = {}
        for row_idx in range(1, min(max_scan_rows + 1, sheet.max_row + 1)):
            values = [sheet.cell(row_idx, col).value for col in range(1, min(sheet.max_column + 1, 30))]
            scores = {
                "field_name": self._score_row_as_role(values, "field_name"),
                "type": self._score_row_as_role(values, "type"),
                "description": self._score_row_as_role(values, "description"),
                "export_tag": self._score_row_as_role(values, "export_tag"),
            }
            best_role = max(scores, key=scores.get)
            best_score = scores[best_role]
            results[row_idx] = {"scores": scores, "best_role": best_role, "best_score": best_score}
        return results

    def _score_row_as_role(self, values: list, role: str) -> float:
        """对一行数据评分它属于某种角色的可能性，返回 0.0~1.0"""
        non_null = [v for v in values if v is not None]
        if not non_null:
            return 0.0
        str_values = [str(v).strip() for v in non_null]
        total = len(str_values)
        if total == 0:
            return 0.0

        if role == "field_name":
            pattern = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
            matches = sum(1 for v in str_values if pattern.match(v) and len(v) > 1)
            return matches / total

        elif role == "type":
            known_types = {'int', 'string', 'str', 'bool', 'boolean', 'float', 'double', 'long', 'object', 'array', 'list', 'dict'}
            matches = sum(1 for v in str_values
                          if v.lower() in known_types
                          or re.match(r'^E[A-Z]', v)
                          or v.endswith('[]')
                          or v.strip().startswith('#')
                          or re.match(r'^\{.*\}$', v))  # 复合类型如 {Type;Num1;Num2}
            return matches / total

        elif role == "description":
            cjk = re.compile(r'[\u4e00-\u9fff]')
            matches = sum(1 for v in str_values if cjk.search(v))
            return matches / total

        elif role == "export_tag":
            matches = sum(1 for v in str_values if 'server' in v.lower() or 'client' in v.lower())
            return matches / total

        return 0.0

    def _vote_on_roles(self, per_file_results: list) -> dict:
        """跨文件投票确定每行的标准角色"""
        role_votes = {}  # {row_idx: {role: count}}
        for file_result in per_file_results:
            for row_idx, info in file_result["roles"].items():
                if row_idx not in role_votes:
                    role_votes[row_idx] = Counter()
                best = info["best_role"]
                if info["best_score"] > 0.3:  # 只计有效评分
                    role_votes[row_idx][best] += 1

        # 确定每行角色（避免冲突）
        row_roles = {}
        used_roles = set()
        total_files = len(per_file_results)
        for row_idx in sorted(role_votes.keys()):
            votes = role_votes[row_idx]
            for role, count in votes.most_common():
                if role not in used_roles and count > total_files * 0.3:
                    row_roles[row_idx] = role
                    used_roles.add(role)
                    break

        # 确定 total_rows 和 data_start_row
        max_header_row = max(row_roles.keys()) if row_roles else 4
        field_name_row = next((r for r, role in row_roles.items() if role == "field_name"), 3)

        return {
            "total_rows": max_header_row,
            "data_start_row": max_header_row + 1,
            "rows": row_roles,
            "field_name_row": field_name_row
        }

    def _detect_reference_patterns(self, scan_result: dict) -> dict:
        """从字段名中发现引用模式"""
        all_fields = []
        for file_info in scan_result["files"]:
            for sheet_info in file_info["sheets"]:
                for header in sheet_info["headers"]:
                    all_fields.append(header["name"])

        suffix_counter = Counter()
        for field in all_fields:
            for suffix in ["Id", "Ids", "ID", "_id", "_ref", "Ref"]:
                if field.endswith(suffix) and len(field) > len(suffix):
                    suffix_counter[suffix] += 1

        naming_patterns = []
        # 只记录出现3次以上的模式
        for suffix, count in suffix_counter.most_common(10):
            if count < 3:
                continue
            if suffix == "Id":
                naming_patterns.append({"pattern": "^(.+)Id$", "target": "{1}.Id", "type": "single_reference", "frequency": count})
            elif suffix == "Ids":
                naming_patterns.append({"pattern": "^(.+)Ids$", "target": "{1}.Id", "type": "array_reference", "frequency": count})
            else:
                naming_patterns.append({"pattern": f"^(.+){re.escape(suffix)}$", "target": "{1}", "type": "candidate_reference", "frequency": count, "confidence": "low"})

        return {
            "naming_patterns": naming_patterns,
            "type_patterns": [
                {"match": "^E[A-Z]", "reference_type": "enum"},
                {"match": "^#", "behavior": "comment_column"},
                {"match": "\\[\\]$", "reference_type": "array"}
            ],
            "discovered_patterns": []
        }

    def _detect_sheet_naming(self, scan_result: dict) -> dict:
        """检测 Sheet 命名约定"""
        separators = Counter()
        for file_info in scan_result["files"]:
            for sheet_info in file_info["sheets"]:
                name = sheet_info["name"]
                if '|' in name:
                    separators['|'] += 1
                elif '-' in name and any(c.isascii() and c.isalpha() for c in name.split('-')[-1]):
                    separators['-'] += 1

        if not separators:
            return {"separator": None, "use_english_part": False}

        sep = separators.most_common(1)[0][0]
        return {"separator": sep, "use_english_part": True}

    def _detect_overrides(self, per_file_results: list, header_config: dict) -> dict:
        """为异常文件生成覆盖配置"""
        overrides = {}
        field_name_row = header_config.get("field_name_row", 3)
        total_rows = header_config.get("total_rows", 4)

        for file_result in per_file_results:
            roles = file_result["roles"]
            file_field_row = next((r for r, info in roles.items()
                                   if info["best_role"] == "field_name" and info["best_score"] > 0.3), None)
            if file_field_row and file_field_row != field_name_row:
                overrides[file_result["file"]] = {
                    "field_name_row": file_field_row,
                    "total_rows": file_field_row + 1,
                    "data_start_row": file_field_row + 2
                }

        return overrides

    def _calculate_confidence(self, header_config: dict, per_file_results: list) -> float:
        """计算整体检测置信度"""
        if not per_file_results:
            return 0.0

        field_row = header_config.get("field_name_row", 3)
        consistent = 0
        for file_result in per_file_results:
            roles = file_result["roles"]
            if field_row in roles and roles[field_row]["best_role"] == "field_name":
                consistent += 1

        return consistent / len(per_file_results)


def main():
    """主函数"""
    import sys

    if len(sys.argv) < 2:
        print("用法:")
        print("  python analyzer.py <配表目录>          # 分析配表")
        print("  python analyzer.py --init <配表目录>    # 自适应初始化")
        sys.exit(1)

    # 处理 --init 参数
    if sys.argv[1] == "--init":
        if len(sys.argv) < 3:
            print("用法: python analyzer.py --init <配表目录>")
            sys.exit(1)
        config_dir = sys.argv[2]
        analyzer = ConfigAnalyzer(config_dir)

        print("=" * 60)
        print("自适应初始化 - 检测项目配表格式")
        print("=" * 60)

        config = analyzer.auto_detect_header_format(sample_count=10)

        print("\n检测结果:")
        print(f"  表头行数: {config['header']['total_rows']}")
        print(f"  字段名行: {config['header']['field_name_row']}")
        print(f"  数据起始: 第 {config['header']['data_start_row']} 行")
        print(f"  置信度:   {config['header'].get('confidence', 0):.0%}")
        print(f"  行角色:   {config['header'].get('rows', {})}")

        if config['references']['naming_patterns']:
            print("\n  引用模式:")
            for p in config['references']['naming_patterns']:
                print(f"    {p['pattern']} -> {p.get('target', '?')} (出现 {p.get('frequency', '?')} 次)")

        overrides = config['header'].get('overrides', {})
        if overrides:
            print(f"\n  异常文件覆盖: {len(overrides)} 个")
            for fname, oc in overrides.items():
                print(f"    {fname}: 字段名在第 {oc['field_name_row']} 行")

        analyzer.save_project_config(config)
        return

    config_dir = sys.argv[1]
    analyzer = ConfigAnalyzer(config_dir)

    # 如果无项目配置，提示初始化
    if not analyzer.project_config:
        print("提示: 未检测到 .gameconfig.yaml 配置文件")
        print("  建议运行: python analyzer.py --init <配表目录>")
        print("  当前使用默认格式假设\n")

    # 扫描配表
    print("扫描配表...")
    scan_result = analyzer.scan_directory()
    print(f"找到 {scan_result['total_count']} 个文件, {scan_result['sheets_count']} 个 Sheet")

    # 分析关系
    print("分析关系...")
    relations = analyzer.analyze_relations(scan_result)
    print(f"找到 {len(relations)} 个引用关系")

    # 提取约束
    print("提取约束...")
    constraints = analyzer.extract_constraints(scan_result)
    print(f"找到 {len(constraints)} 个约束规则")

    # 生成图表
    mermaid = analyzer.generate_mermaid_graph(relations)
    print("\nMermaid 图表:")
    print(mermaid)

    # 保存结果
    output = {
        "scan_result": scan_result,
        "relations": relations,
        "constraints": constraints,
        "mermaid_graph": mermaid
    }

    output_file = "config_analysis_result.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n结果已保存到: {output_file}")


if __name__ == "__main__":
    main()
