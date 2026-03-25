#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配表全文搜索脚本

跨所有配表搜索特定值
"""

import openpyxl
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Any, Optional
from collections import defaultdict


class ConfigSearcher:
    """配表全文搜索器"""

    def __init__(self, config_dir: str = None):
        self.config_dir = Path(config_dir) if config_dir else Path.cwd()
        self.index = {}
        self.tables = {}

    def build_index(self):
        """构建搜索索引"""
        print("扫描配表构建索引...")

        for file_path in self.config_dir.rglob("*.xlsx"):
            if "~$" in file_path.name:
                continue

            try:
                self._index_file(file_path)
            except Exception as e:
                print(f"警告: 无法索引 {file_path}: {e}")

        print(f"索引完成: {len(self.tables)} 个表")

    def _index_file(self, file_path: Path):
        """索引单个文件"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet in wb.worksheets:
            table_key = f"{file_path.name}/{sheet.title}"
            self.tables[table_key] = {
                "file": file_path.name,
                "sheet": sheet.title,
                "path": str(file_path)
            }

            # 索引数据
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_idx, value in enumerate(row, start=1):
                    if value is not None:
                        value_str = str(value)
                        # 索引值
                        if value_str not in self.index:
                            self.index[value_str] = []

                        self.index[value_str].append({
                            "table": table_key,
                            "row": row_idx,
                            "col": col_idx,
                            "value": value_str
                        })

    def search(self, query: str, search_type: str = "fuzzy", field: str = None) -> Dict[str, Any]:
        """执行搜索"""
        results = {
            "query": query,
            "type": search_type,
            "total_matches": 0,
            "matches": []
        }

        if search_type == "exact":
            matches = self._exact_search(query, field)
        elif search_type == "fuzzy":
            matches = self._fuzzy_search(query, field)
        elif search_type == "regex":
            matches = self._regex_search(query, field)
        elif search_type == "field":
            matches = self._field_search(query)
        else:
            matches = []

        # 按表分组
        grouped = defaultdict(list)
        for match in matches:
            table_key = match["table"]
            grouped[table_key].append(match)

        # 转换为列表
        for table_key, table_matches in grouped.items():
            table_info = self.tables.get(table_key, {})
            results["matches"].append({
                "table": table_key,
                "file": table_info.get("file", ""),
                "sheet": table_info.get("sheet", ""),
                "count": len(table_matches),
                "matches": table_matches[:50]  # 限制每个表最多返回 50 个结果
            })

        results["total_matches"] = sum(m["count"] for m in results["matches"])

        return results

    def _exact_search(self, query: str, field: str = None) -> List[Dict]:
        """精确搜索"""
        if query in self.index:
            matches = self.index[query]
            if field:
                return [m for m in matches if self._get_field_name(m) == field]
            return matches[:]
        return []

    def _fuzzy_search(self, query: str, field: str = None) -> List[Dict]:
        """模糊搜索"""
        query_lower = query.lower()
        matches = []

        for value, locations in self.index.items():
            if query_lower in value.lower():
                if field:
                    matches.extend([m for m in locations if self._get_field_name(m) == field])
                else:
                    matches.extend(locations)

        return matches

    def _regex_search(self, query: str, field: str = None) -> List[Dict]:
        """正则搜索"""
        try:
            pattern = re.compile(query)
        except re.error:
            return []

        matches = []
        for value, locations in self.index.items():
            if pattern.search(value):
                if field:
                    matches.extend([m for m in locations if self._get_field_name(m) == field])
                else:
                    matches.extend(locations)

        return matches

    def _field_search(self, query: str) -> List[Dict]:
        """字段限定搜索 field:value"""
        try:
            field, value = query.split(":", 1)
        except ValueError:
            return []

        matches = []
        for value_key, locations in self.index.items():
            if value.lower() in value_key.lower():
                matches.extend([m for m in locations if self._get_field_name(m) == field])

        return matches

    def _get_field_name(self, match: Dict) -> Optional[str]:
        """获取匹配位置的字段名"""
        table_key = match["table"]
        table_info = self.tables.get(table_key, {})
        file_path = Path(table_info.get("path", ""))

        if not file_path.exists():
            return None

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[table_info["sheet"]]

            # 获取表头（假设第 4 行是字段名）
            header_row = 3
            field_name = sheet.cell(header_row, match["col"]).value
            return str(field_name) if field_name else None
        except:
            return None

    def generate_report(self, result: Dict) -> str:
        """生成搜索报告"""
        lines = [
            "# 配表搜索结果",
            f"",
            f"**查询**: `{result['query']}`",
            f"**类型**: {result['type']}",
            f"**匹配数**: {result['total_matches']}",
            f"**时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f""
        ]

        if not result["matches"]:
            lines.extend([
                f"## 无匹配",
                f"",
                f"没有找到匹配 '{result['query']}' 的记录",
                f""
            ])
        else:
            lines.extend([
                f"## 匹配结果",
                f""
            ])

            for match_group in result["matches"]:
                lines.extend([
                    f"### {match_group['file']}",
                    f"",
                    f"**Sheet**: {match_group['sheet']}",
                    f"**匹配数**: {match_group['count']}",
                    f""
                ])

                if match_group["matches"]:
                    lines.extend([
                        f"| 行 | 列 | 字段 | 值 |",
                        f"|----|----|------|-----|"
                    ])

                    for m in match_group["matches"][:20]:
                        field_name = self._get_field_name(m)
                        value_display = m["value"][:50] + "..." if len(m["value"]) > 50 else m["value"]
                        lines.append(f"| {m['row']} | {m['col']} | {field_name or '-'} | `{value_display}` |")

                    if match_group["count"] > 20:
                        lines.append(f"| ... | ... | ... | 还有 {match_group['count'] - 20} 个匹配 |")

                lines.append("")

        return "\n".join(lines)

    def save_index(self, index_file: str):
        """保存索引"""
        data = {
            "timestamp": datetime.now().isoformat(),
            "tables": self.tables,
            "index": {k: v[:100] for k, v in list(self.index.items())[:1000]}  # 限制大小
        }

        with open(index_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"索引已保存到: {index_file}")

    def load_index(self, index_file: str):
        """加载索引"""
        with open(index_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            self.tables = data["tables"]
            # 注意：完整的索引可能很大，这里简化处理
            self.index = {}
            print(f"索引已加载: {len(self.tables)} 个表")


def main():
    """主函数"""
    import sys

    if len(sys.argv) < 2:
        print("用法:")
        print("  python search.py <配表目录> <查询> [类型]")
        print("")
        print("查询类型:")
        print("  exact  - 精确匹配 (默认)")
        print("  fuzzy  - 模糊匹配")
        print("  regex  - 正则表达式")
        print("  field  - 字段限定 (格式: field:value)")
        print("")
        print("示例:")
        print("  python search.py ./configs 关卡")
        print("  python search.py ./configs '^H\\d+$' regex")
        print("  python search.py ./configs 'Name:关羽' field")
        sys.exit(1)

    config_dir = sys.argv[1]
    query = sys.argv[2] if len(sys.argv) > 2 else ""
    search_type = sys.argv[3] if len(sys.argv) > 3 else "fuzzy"

    searcher = ConfigSearcher(config_dir)

    # 检查是否有已保存的索引
    index_file = Path(config_dir) / ".search_index.json"
    if index_file.exists():
        print(f"加载已保存的索引...")
        searcher.load_index(str(index_file))
    else:
        searcher.build_index()
        searcher.save_index(str(index_file))

    # 执行搜索
    print(f"搜索: {query} (类型: {search_type})")
    result = searcher.search(query, search_type)

    # 保存结果
    output_file = "search_result.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"找到 {result['total_matches']} 个匹配")
    print(f"结果已保存到: {output_file}")

    # 生成报告
    report = searcher.generate_report(result)
    report_file = output_file.replace('.json', '_report.md')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"报告已保存到: {report_file}")

    # 打印摘要
    if result["matches"]:
        print()
        print("匹配摘要:")
        for match_group in result["matches"][:5]:
            print(f"  {match_group['file']}.{match_group['sheet']}: {match_group['count']} 个匹配")


if __name__ == "__main__":
    main()
