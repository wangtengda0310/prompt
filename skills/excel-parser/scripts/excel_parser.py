#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 配表解析脚本 - 直接读取游戏配表 Excel 文件
支持名将杀项目统一的 4 行表头结构
"""

import sys
import json
import argparse
import io
from pathlib import Path

# 强制 UTF-8 输出（解决 Windows Git Bash 中文乱码）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl。运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 名将杀项目表头结构常量
HEADER_ROWS = 4  # 前4行是表头
CHS_NAME_ROW = 0  # 第1行: 中文名称
TYPE_ROW = 1      # 第2行: 字段类型
FIELD_ROW = 2     # 第3行: 字段名
EXPORT_ROW = 3    # 第4行: 导出标识
DATA_START = 5    # 第5行开始是数据


def load_excel(file_path: str, sheet_name: str = None):
    """加载 Excel 文件，返回工作簿对象"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name:
        # 尝试直接匹配
        if sheet_name in wb.sheetnames:
            return wb, wb[sheet_name]
        # 尝试后缀匹配（支持 "中文|英文" 格式）
        for name in wb.sheetnames:
            if '|' in name:
                parts = name.split('|')
                if sheet_name in parts:
                    return wb, wb[name]
            if sheet_name in name:
                return wb, wb[name]
        available = ', '.join(wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' 不存在。可用: {available}")
    return wb, wb[wb.sheetnames[0]]


def get_sheet_info(file_path: str):
    """获取 Excel 文件的所有 Sheet 信息"""
    wb, _ = load_excel(file_path)
    result = {
        "filePath": file_path,
        "sheets": []
    }
    for name in wb.sheetnames:
        sheet = wb[name]
        result["sheets"].append({
            "name": name,
            "maxRow": sheet.max_row,
            "maxCol": sheet.max_column
        })
    return result


def get_column_info(file_path: str, sheet_name: str):
    """获取 Sheet 的列定义信息"""
    wb, sheet = load_excel(file_path, sheet_name)
    columns = []

    for col_idx in range(1, sheet.max_column + 1):
        chs_name = sheet.cell(1, col_idx).value
        field_type = sheet.cell(2, col_idx).value
        field_name = sheet.cell(3, col_idx).value
        export_tag = sheet.cell(4, col_idx).value

        status = "NORMAL"
        if field_type == "#":
            status = "COMMENT"
        elif field_type and str(field_type).startswith("E#"):
            status = "ENUM"
        elif not field_name:
            status = "EMPTY"

        columns.append({
            "index": col_idx,
            "chsName": chs_name or "",
            "fieldType": field_type or "",
            "fieldName": field_name or "",
            "exportTag": export_tag or "",
            "status": status
        })

    return {
        "filePath": file_path,
        "sheetName": sheet_name,
        "columns": columns
    }


def preview_sheet(file_path: str, sheet_name: str, rows: int = 10):
    """预览 Sheet 前 N 行数据"""
    wb, sheet = load_excel(file_path, sheet_name)

    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        chs_name = sheet.cell(1, col_idx).value
        headers.append(str(chs_name) if chs_name else f"Col{col_idx}")

    data_rows = []
    end_row = min(DATA_START + rows - 1, sheet.max_row)
    for row_idx in range(DATA_START, end_row + 1):
        row_data = []
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row_idx, col_idx).value
            row_data.append(str(val) if val is not None else "")
        data_rows.append(row_data)

    return {
        "filePath": file_path,
        "sheetName": sheet_name,
        "headers": headers,
        "dataRows": data_rows,
        "totalRows": sheet.max_row - HEADER_ROWS,
        "returnedRows": len(data_rows)
    }


def query_range(file_path: str, sheet_name: str, start_row: int, end_row: int = None, include_header: bool = False):
    """查询指定行范围的数据"""
    wb, sheet = load_excel(file_path, sheet_name)

    # start_row 是数据行号（从1开始），转换为 Excel 行号
    excel_start = DATA_START + start_row - 1
    excel_end = DATA_START + end_row - 1 if end_row else sheet.max_row

    if excel_start > sheet.max_row:
        return {"error": f"起始行 {start_row} 超出数据范围"}

    excel_end = min(excel_end, sheet.max_row)

    result = {
        "filePath": file_path,
        "sheetName": sheet_name,
        "startRow": start_row,
        "endRow": start_row + (excel_end - excel_start),
        "totalRows": sheet.max_row - HEADER_ROWS,
        "headers": [],
        "dataRows": []
    }

    if include_header:
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            chs_name = sheet.cell(1, col_idx).value
            field_name = sheet.cell(3, col_idx).value
            headers.append(str(chs_name) if chs_name else str(field_name) or f"Col{col_idx}")
        result["headers"] = headers

    for row_idx in range(excel_start, excel_end + 1):
        row_data = []
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row_idx, col_idx).value
            row_data.append(str(val) if val is not None else "")
        result["dataRows"].append(row_data)

    return result


def filter_data(file_path: str, sheet_name: str, conditions: list, include_header: bool = False):
    """根据条件过滤数据"""
    wb, sheet = load_excel(file_path, sheet_name)

    # 构建列名到索引的映射
    col_map = {}
    for col_idx in range(1, sheet.max_column + 1):
        chs_name = sheet.cell(1, col_idx).value
        field_name = sheet.cell(3, col_idx).value
        if chs_name:
            col_map[str(chs_name)] = col_idx
        if field_name:
            col_map[str(field_name)] = col_idx

    result = {
        "filePath": file_path,
        "sheetName": sheet_name,
        "totalRows": sheet.max_row - HEADER_ROWS,
        "matchedRows": 0,
        "headers": [],
        "dataRows": []
    }

    if include_header:
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            chs_name = sheet.cell(1, col_idx).value
            field_name = sheet.cell(3, col_idx).value
            headers.append(str(chs_name) if chs_name else str(field_name) or f"Col{col_idx}")
        result["headers"] = headers

    for row_idx in range(DATA_START, sheet.max_row + 1):
        match = True
        for cond in conditions:
            col_name = cond.get("columnName", "")
            value = cond.get("value", "")
            operator = cond.get("operator", "eq")

            if col_name not in col_map:
                match = False
                break

            col_idx = col_map[col_name]
            cell_val = sheet.cell(row_idx, col_idx).value
            cell_str = str(cell_val) if cell_val is not None else ""

            if operator == "eq":
                match = cell_str == value
            elif operator == "neq":
                match = cell_str != value
            elif operator == "contains":
                match = value in cell_str
            elif operator == "startsWith":
                match = cell_str.startswith(value)
            elif operator == "endsWith":
                match = cell_str.endswith(value)

            if not match:
                break

        if match:
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                val = sheet.cell(row_idx, col_idx).value
                row_data.append(str(val) if val is not None else "")
            result["dataRows"].append(row_data)

    result["matchedRows"] = len(result["dataRows"])
    return result


def scan_excel_dir(dir_path: str):
    """扫描目录下的所有 Excel 文件"""
    path = Path(dir_path)
    files = []
    for f in sorted(path.glob("*.xlsx")):
        if f.name.startswith("~$"):  # 跳过临时文件
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            files.append({
                "fileName": f.name,
                "filePath": str(f),
                "sheets": sheets
            })
        except Exception as e:
            files.append({
                "fileName": f.name,
                "filePath": str(f),
                "error": str(e)
            })

    return {"dirPath": dir_path, "files": files}


def main():
    parser = argparse.ArgumentParser(description="Excel 配表解析工具")
    parser.add_argument("action", choices=[
        "scan", "sheets", "columns", "preview", "query", "filter"
    ], help="操作类型")
    parser.add_argument("--file", "-f", help="Excel 文件路径")
    parser.add_argument("--dir", "-d", help="Excel 目录路径")
    parser.add_argument("--sheet", "-s", help="Sheet 名称")
    parser.add_argument("--rows", "-r", type=int, default=10, help="预览行数")
    parser.add_argument("--start", type=int, default=1, help="起始数据行号")
    parser.add_argument("--end", type=int, help="结束数据行号")
    parser.add_argument("--header", action="store_true", help="包含表头")
    parser.add_argument("--conditions", "-c", help="过滤条件 JSON")

    args = parser.parse_args()

    try:
        if args.action == "scan":
            if not args.dir:
                print("错误: scan 操作需要 --dir 参数", file=sys.stderr)
                sys.exit(1)
            result = scan_excel_dir(args.dir)

        elif args.action == "sheets":
            if not args.file:
                print("错误: sheets 操作需要 --file 参数", file=sys.stderr)
                sys.exit(1)
            result = get_sheet_info(args.file)

        elif args.action == "columns":
            if not args.file or not args.sheet:
                print("错误: columns 操作需要 --file 和 --sheet 参数", file=sys.stderr)
                sys.exit(1)
            result = get_column_info(args.file, args.sheet)

        elif args.action == "preview":
            if not args.file or not args.sheet:
                print("错误: preview 操作需要 --file 和 --sheet 参数", file=sys.stderr)
                sys.exit(1)
            result = preview_sheet(args.file, args.sheet, args.rows)

        elif args.action == "query":
            if not args.file or not args.sheet:
                print("错误: query 操作需要 --file 和 --sheet 参数", file=sys.stderr)
                sys.exit(1)
            result = query_range(args.file, args.sheet, args.start, args.end, args.header)

        elif args.action == "filter":
            if not args.file or not args.sheet or not args.conditions:
                print("错误: filter 操作需要 --file, --sheet 和 --conditions 参数", file=sys.stderr)
                sys.exit(1)
            conditions = json.loads(args.conditions)
            result = filter_data(args.file, args.sheet, conditions, args.header)

        print(json.dumps(result, ensure_ascii=False, indent=2))

    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
